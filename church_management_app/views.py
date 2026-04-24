# views.py - Vues Django traditionnelles pour le rendu côté serveur
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from datetime import timedelta

from .models import (
    Member, Family, HomeGroup, Department, Ministry, Event, Attendance,
    EvangelismActivity, TrainingEvent, MarriageRecord,
    FinancialCategory, FinancialTransaction, Announcement, Document, LogisticsItem,
    LogisticsCategory, LogisticsCondition,
    ChurchBiography, Contact, ChurchSettings, EventAttendanceAggregate,
    BaptismEvent, BaptismCandidate, AuditLogEntry, ApprovalRequest, Notification,
    ChurchActivity, ChurchService
)
from .forms import (
    MemberForm, FamilyForm, HomeGroupForm, DepartmentForm, MinistryForm,
    EventForm, AttendanceForm, EvangelismActivityForm, TrainingEventForm, MarriageRecordForm,
    FinancialCategoryForm, FinancialTransactionForm, AnnouncementForm, DocumentForm, DocumentEditForm, LogisticsItemForm,
    BaptismCandidateForm,
    LoginForm, UserCreateForm, UserUpdateForm, ProfileUpdateForm, PasswordChangeCustomForm,
    EventAttendanceAggregateForm, ChurchSettingsForm, ChurchActivityForm, ChurchBiographyForm, ChurchServiceForm
)
from .permissions import (
    role_required, admin_required, finance_required, pastor_required,
    member_management_required, can_access_menu, can_create, can_edit, can_delete
)

User = get_user_model()


# ============================================================
# Authentification
# ============================================================

def login_view(request):
    """Page de connexion avec authentification Django"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    # Récupérer les paramètres de l'église
    try:
        church_settings = ChurchSettings.objects.first()
    except:
        church_settings = None
    
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Bienvenue, {user.first_name or username}!')
                return redirect('dashboard')
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form, 'church_settings': church_settings})


def logout_view(request):
    """Déconnexion"""
    logout(request)
    messages.success(request, 'Vous avez été déconnecté.')
    return redirect('home')


# ============================================================
# Pages publiques
# ============================================================

def index(request):
    """Page d'accueil publique"""
    context = {
        'active_page': 'home',
        'upcoming_events': Event.objects.filter(
            date__gte=timezone.now().date(),
            is_published=True
        ).order_by('date', 'time')[:5],
        'activities': ChurchActivity.objects.filter(is_active=True).order_by('order', 'title'),
        'church_services': ChurchService.objects.filter(is_active=True).order_by('order', 'day', 'time'),
    }
    return render(request, 'index.html', context)


def public_about(request):
    """Page À propos"""
    context = {
        'active_page': 'about',
        'church_biography': ChurchBiography.objects.filter(is_active=True).first(),
    }
    return render(request, 'public-about.html', context)


def contact(request):
    """Page Contact"""
    if request.method == 'POST':
        from .forms import ContactForm
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Votre message a été envoyé avec succès!')
            return redirect('contact')
    else:
        from .forms import ContactForm
        form = ContactForm()
    
    return render(request, 'contact.html', {'form': form, 'active_page': 'contact'})


def public_events(request):
    """Page publique des événements"""
    today = timezone.now().date()
    
    events = Event.objects.filter(
        is_published=True
    ).order_by('date', 'time')
    
    upcoming_events = events.filter(date__gte=today)
    past_events = events.filter(date__lt=today)
    
    event_type = request.GET.get('type', '')
    if event_type:
        upcoming_events = upcoming_events.filter(event_type=event_type)
        past_events = past_events.filter(event_type=event_type)
    
    return render(request, 'public-events.html', {
        'active_page': 'events',
        'upcoming_events': upcoming_events,
        'past_events': past_events,
        'event_type': event_type,
    })


# ============================================================
# Tableau de bord
# ============================================================

@login_required
def dashboard(request):
    """Tableau de bord admin avec statistiques complètes"""
    user = request.user
    
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_start = today.replace(day=1)
    
    # Statistiques membres
    total_members = Member.objects.count()
    new_members_week = Member.objects.filter(created_at__date__gte=week_ago).count()
    new_members_month = Member.objects.filter(created_at__date__gte=month_start).count()
    
    # Statistiques événements
    upcoming_events = Event.objects.filter(date__gte=today, is_published=True).count()
    events_this_week = Event.objects.filter(date__gte=week_ago, date__lte=today).count()
    events_this_month = Event.objects.filter(date__month=today.month, date__year=today.year).count()
    
    # Événements à venir (liste)
    upcoming_events_list = Event.objects.filter(
        date__gte=today, is_published=True
    ).order_by('date', 'time')[:5]
    
    # Derniers événements passés
    recent_events = Event.objects.filter(
        date__lt=today
    ).order_by('-date')[:3]
    
    # Statistiques finances
    total_in_month = FinancialTransaction.objects.filter(
        direction='in', date__gte=month_start
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    total_out_month = FinancialTransaction.objects.filter(
        direction='out', date__gte=month_start
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Statistiques présences
    attendance_week = Attendance.objects.filter(
        event__date__gte=week_ago, attended=True
    ).count()
    
    # Annonces actives
    active_announcements = Announcement.objects.filter(is_active=True).count()
    recent_announcements = Announcement.objects.filter(
        is_active=True
    ).order_by('-created_at')[:5]
    
    # Anniversaires du mois
    from django.db.models.functions import ExtractMonth, ExtractDay
    birthdays_this_month = Member.objects.annotate(
        birth_month=ExtractMonth('birth_date'),
        birth_day=ExtractDay('birth_date')
    ).filter(birth_month=today.month).order_by('birth_day')
    
    # Paramètres de l'église
    try:
        church_settings = ChurchSettings.objects.first()
    except:
        church_settings = None
    
    context = {
        'user': user,
        'today': today,
        # Membres
        'total_members': total_members,
        'new_members_week': new_members_week,
        'new_members_month': new_members_month,
        # Événements
        'upcoming_events': upcoming_events,
        'events_this_week': events_this_week,
        'events_this_month': events_this_month,
        'upcoming_events_list': upcoming_events_list,
        'recent_events': recent_events,
        # Finances
        'total_in_month': total_in_month,
        'total_out_month': total_out_month,
        'balance_month': total_in_month - total_out_month,
        # Présences
        'attendance_week': attendance_week,
        # Annonces
        'active_announcements': active_announcements,
        'recent_announcements': recent_announcements,
        # Anniversaires
        'birthdays_this_month': birthdays_this_month,
        # Paramètres de l'église
        'church_settings': church_settings,
    }
    return render(request, 'dashboard/dashboard.html', context)


# ============================================================
# Membres - CRUD
# ============================================================

@login_required
def member_list(request):
    """Liste des membres avec pagination, recherche et filtres"""
    members = Member.objects.select_related('user').all()
    
    # Recherche
    search = request.GET.get('q', '')
    if search:
        members = members.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(member_number__icontains=search)
        )
    
    # Filtre par statut
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        members = members.filter(is_active=True)
    elif status_filter == 'inactive':
        members = members.filter(is_active=False)
    elif status_filter == 'visitor':
        members = members.filter(user__isnull=True)
    
    # Statistiques
    from django.utils import timezone
    from datetime import timedelta
    
    total_members = Member.objects.count()
    active_members = Member.objects.filter(is_active=True).count()
    new_members = Member.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=30)
    ).count()
    married_members = Member.objects.filter(marital_status='married').count()
    
    # Pagination
    paginator = Paginator(members, 20)  # 20 membres par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dashboard/members.html', {
        'members': page_obj,
        'search': search,
        'status_filter': status_filter,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_members': total_members,
        'active_members': active_members,
        'new_members': new_members,
        'married_members': married_members,
    })


@login_required
def member_detail(request, pk):
    """Détail d'un membre"""
    member = get_object_or_404(Member, pk=pk)
    attendance_history = Attendance.objects.filter(
        member=member
    ).order_by('-event__date')[:10]
    
    return render(request, 'dashboard/member-detail.html', {
        'member': member,
        'attendance_history': attendance_history
    })


@login_required
@member_management_required
def member_create(request):
    """Créer un nouveau membre"""
    if not can_create(request.user):
        messages.error(request, "Vous n'avez pas la permission de créer des membres.")
        return redirect('member-list')
    
    if request.method == 'POST':
        form = MemberForm(request.POST, request.FILES)
        if form.is_valid():
            member = form.save()
            messages.success(request, f'Membre {member.user.first_name} {member.user.last_name} créé avec succès!')
            return redirect('member-list')
    else:
        form = MemberForm()
    
    return render(request, 'dashboard/member-form.html', {
        'form': form,
        'title': 'Ajouter un Membre',
        'action': 'Créer'
    })


@login_required
@member_management_required
def member_edit(request, pk):
    """Modifier un membre"""
    if not can_edit(request.user):
        messages.error(request, "Vous n'avez pas la permission de modifier des membres.")
        return redirect('member-detail', pk=pk)
    
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        form = MemberForm(request.POST, request.FILES, instance=member)
        if form.is_valid():
            form.save()
            messages.success(request, f'Membre {member.user.first_name} {member.user.last_name} modifié avec succès!')
            return redirect('member-list')
    else:
        form = MemberForm(instance=member)
    
    return render(request, 'dashboard/member-form.html', {
        'form': form,
        'member': member,
        'title': 'Modifier le Membre',
        'action': 'Modifier'
    })


@login_required
@member_management_required
def member_delete(request, pk):
    """Supprimer un membre"""
    if not can_delete(request.user):
        messages.error(request, "Vous n'avez pas la permission de supprimer des membres.")
        return redirect('member-detail', pk=pk)
    
    member = get_object_or_404(Member, pk=pk)
    
    if request.method == 'POST':
        member.delete()
        messages.success(request, 'Membre supprimé avec succès!')
        return redirect('member-list')
    
    return render(request, 'dashboard/members.html', {'member': member, 'delete_confirm': True})


@login_required
def member_print_list(request):
    """Imprimer la liste des membres (PDF) - Version professionnelle A4"""
    from django.http import HttpResponse
    from django.conf import settings
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.pdfgen import canvas
    from io import BytesIO
    from datetime import datetime
    import os

    # Récupérer les paramètres de l'église
    church_settings = ChurchSettings.objects.first()
    church_name = church_settings.church_name if church_settings else "Consolation et Paix Divine"
    church_slogan = church_settings.church_slogan if church_settings else ""
    church_address = church_settings.address if church_settings else ""
    church_city = church_settings.city if church_settings else ""
    church_phone = church_settings.phone_primary if church_settings else ""
    church_email = church_settings.email_primary if church_settings else ""
    logo_path = church_settings.logo.path if church_settings and church_settings.logo else None

    # Récupérer les membres avec filtres optionnels
    members = Member.objects.select_related('user', 'department', 'ministry', 'home_group').all()

    # Appliquer les filtres si présents
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        members = members.filter(is_active=True)
    elif status_filter == 'inactive':
        members = members.filter(is_active=False)

    department_filter = request.GET.get('department', '')
    if department_filter:
        members = members.filter(department_id=department_filter)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="liste_membres_{datetime.now().strftime("%Y%m%d")}.pdf"'

    buffer = BytesIO()

    # Configuration du document A4 avec marges
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=25*mm,
        bottomMargin=20*mm
    )

    # Styles personnalisés
    styles = getSampleStyleSheet()

    # Style pour l'en-tête de l'église
    church_name_style = ParagraphStyle(
        'ChurchName',
        parent=styles['Heading1'],
        fontSize=16,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1a365d'),
        alignment=TA_CENTER,
        spaceAfter=6
    )

    church_slogan_style = ParagraphStyle(
        'ChurchSlogan',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Helvetica-Oblique',
        textColor=colors.HexColor('#4a5568'),
        alignment=TA_CENTER,
        spaceAfter=6
    )

    church_info_style = ParagraphStyle(
        'ChurchInfo',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#2d3748'),
        alignment=TA_CENTER,
        spaceAfter=2
    )

    # Style pour le titre du document
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading2'],
        fontSize=14,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2d3748'),
        alignment=TA_CENTER,
        spaceAfter=12,
        spaceBefore=12
    )

    # Style pour le pied de page
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#718096'),
        alignment=TA_CENTER
    )

    elements = []

    # En-tête du document
    header_data = []

    # Logo si disponible
    if logo_path and os.path.exists(logo_path):
        try:
            img = Image(logo_path, width=2*cm, height=2*cm)
            header_data.append([img, '', ''])
        except:
            pass

    # Informations de l'église
    church_header = [
        Paragraph(church_name, church_name_style),
        Paragraph(church_slogan, church_slogan_style) if church_slogan else '',
        Paragraph(f"{church_address}, {church_city}", church_info_style) if church_address else '',
        Paragraph(f"Tél: {church_phone} | Email: {church_email}", church_info_style) if church_phone or church_email else ''
    ]

    for line in church_header:
        if line:
            elements.append(line)

    # Ligne de séparation
    elements.append(Spacer(1, 0.3*cm))

    # Titre du document
    title_text = "REGISTRE DES MEMBRES"
    if status_filter == 'active':
        title_text = "LISTE DES MEMBRES ACTIFS"
    elif status_filter == 'inactive':
        title_text = "LISTE DES MEMBRES INACTIFS"

    elements.append(Paragraph(title_text, title_style))
    elements.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", church_info_style))
    elements.append(Spacer(1, 0.5*cm))

    # Préparation des données du tableau
    headers = ['N°', 'Matricule', 'Nom & Prénoms', 'Téléphone', 'Email', 'Genre', 'Département', 'Statut']
    data = [headers]

    for idx, member in enumerate(members, 1):
        full_name = f"{member.user.last_name} {member.user.first_name}" if member.user else 'N/A'
        if member.post_name:
            full_name += f" ({member.post_name})"

        data.append([
            str(idx),
            member.member_number or '-',
            full_name,
            member.user.phone if member.user else '-',
            member.user.email if member.user else '-',
            member.get_gender_display() if member.gender else '-',
            member.department.name if member.department else '-',
            'Actif' if member.is_active else 'Inactif'
        ])

    # Création du tableau
    col_widths = [0.8*cm, 2.2*cm, 4.5*cm, 2.5*cm, 3.5*cm, 1.5*cm, 2.5*cm, 1.5*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Style du tableau professionnel
    table_style = TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),

        # Corps du tableau
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (5, 1), (7, -1), 'CENTER'),
        ('ALIGN', (2, 1), (4, -1), 'LEFT'),

        # Alternance de couleurs pour les lignes
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),

        # Bordures
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#1a365d')),

        # Espacement
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ])

    # Alternance de couleurs pour les lignes
    for i in range(1, len(data)):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f7fafc'))

    table.setStyle(table_style)
    elements.append(table)

    # Résumé
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f"Total: {len(members)} membre(s)", church_info_style))

    # Fonction pour ajouter le filigrane et pied de page sur chaque page
    def add_watermark_and_footer(canvas, doc):
        canvas.saveState()

        # Filigrane avec le logo (en arrière-plan)
        if logo_path and os.path.exists(logo_path):
            try:
                # Position centrale avec transparence
                canvas.setFillAlpha(0.08)
                img = Image(logo_path)
                img.drawHeight = 8*cm
                img.drawWidth = 8*cm
                img.wrapOn(canvas, A4[0], A4[1])
                img.drawOn(canvas, (A4[0] - 8*cm) / 2, (A4[1] - 8*cm) / 2)
                canvas.setFillAlpha(1.0)
            except:
                pass

        # Pied de page
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#718096'))

        # Ligne de séparation
        canvas.setStrokeColor(colors.HexColor('#cbd5e0'))
        canvas.line(15*mm, 15*mm, A4[0] - 15*mm, 15*mm)

        # Contact info
        footer_text = f"{church_name}"
        if church_phone:
            footer_text += f" | Tél: {church_phone}"
        if church_email:
            footer_text += f" | Email: {church_email}"

        canvas.drawCentredString(A4[0] / 2, 10*mm, footer_text)

        # Numéro de page
        canvas.drawRightString(A4[0] - 15*mm, 10*mm, f"Page {canvas.getPageNumber()}")

        canvas.restoreState()

    # Construction du PDF avec la fonction de callback pour chaque page
    doc.build(elements, onFirstPage=add_watermark_and_footer, onLaterPages=add_watermark_and_footer)

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


@login_required
def member_print_preview(request):
    """Aperçu avant impression de la liste des membres"""
    members = Member.objects.select_related('user', 'department', 'ministry', 'home_group').all()

    # Appliquer les filtres si présents
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        members = members.filter(is_active=True)
    elif status_filter == 'inactive':
        members = members.filter(is_active=False)
    elif status_filter == 'visitor':
        members = members.filter(user__isnull=True)

    department_filter = request.GET.get('department', '')
    if department_filter:
        members = members.filter(department_id=department_filter)

    search = request.GET.get('search', '')
    if search:
        members = members.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(member_number__icontains=search)
        )

    # Récupérer les paramètres de l'église
    church_settings = ChurchSettings.objects.first()

    context = {
        'members': members,
        'church_settings': church_settings,
        'total_members': members.count(),
        'status_filter': status_filter,
        'department_filter': department_filter,
        'search': search,
        'print_date': timezone.now(),
    }

    return render(request, 'dashboard/members_print_preview.html', context)


@login_required
def member_profile_print(request, pk):
    """Fiche complète du membre pour impression"""
    member = get_object_or_404(
        Member.objects.select_related(
            'user', 'department', 'ministry', 'home_group', 'family'
        ),
        pk=pk
    )

    # Récupérer les paramètres de l'église
    church_settings = ChurchSettings.objects.first()

    context = {
        'member': member,
        'church_settings': church_settings,
        'print_date': timezone.now(),
    }

    return render(request, 'dashboard/member_profile_print.html', context)


@login_required
def member_export(request):
    """Exporter les membres en Excel (xlsx) - Version professionnelle"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from datetime import datetime
    from io import BytesIO

    # Récupérer les paramètres de l'église
    church_settings = ChurchSettings.objects.first()
    church_name = church_settings.church_name if church_settings else "Consolation et Paix Divine"

    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres"

    # Styles
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font = Font(name='Calibri', size=14, bold=True, color="1a365d")
    title_alignment = Alignment(horizontal="center", vertical="center")

    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Bordures
    thin_border = Border(
        left=Side(style='thin', color='cbd5e0'),
        right=Side(style='thin', color='cbd5e0'),
        top=Side(style='thin', color='cbd5e0'),
        bottom=Side(style='thin', color='cbd5e0')
    )

    # En-tête du document (titre)
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"{church_name} - REGISTRE DES MEMBRES"
    title_cell.font = title_font
    title_cell.alignment = title_alignment

    # Sous-titre avec date
    ws.merge_cells('A2:K2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Export généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    subtitle_cell.font = Font(name='Calibri', size=10, italic=True, color="718096")
    subtitle_cell.alignment = title_alignment

    # Ligne vide
    ws.append([])

    # En-têtes des colonnes
    headers = [
        'N°', 'Matricule', 'Nom', 'Prénom', 'Post-nom', 'Téléphone', 'Email',
        'Genre', 'Date de naissance', 'Département', 'Ministère', 'Statut'
    ]

    ws.append(headers)
    header_row = ws[4]  # La ligne 4 contient les headers

    # Appliquer le style aux en-têtes
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Récupérer les membres avec filtres
    members = Member.objects.select_related('user', 'department', 'ministry').all()

    # Appliquer les filtres si présents
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        members = members.filter(is_active=True)
    elif status_filter == 'inactive':
        members = members.filter(is_active=False)

    department_filter = request.GET.get('department', '')
    if department_filter:
        members = members.filter(department_id=department_filter)

    # Ajouter les données
    for idx, member in enumerate(members, 1):
        row_data = [
            idx,
            member.member_number or '',
            member.user.last_name if member.user else '',
            member.user.first_name if member.user else '',
            member.post_name or '',
            member.user.phone if member.user else '',
            member.user.email if member.user else '',
            member.get_gender_display() if member.gender else '',
            member.birth_date.strftime('%d/%m/%Y') if member.birth_date else '',
            member.department.name if member.department else '',
            member.ministry.name if member.ministry else '',
            'Actif' if member.is_active else 'Inactif'
        ]
        ws.append(row_data)

    # Appliquer les styles aux données
    for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row), start=5):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_alignment

            # Alternance de couleurs pour les lignes
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="f7fafc", end_color="f7fafc", fill_type="solid")

            # Alignement centré pour certaines colonnes
            if cell.column in [1, 2, 8, 12]:  # N°, Matricule, Genre, Statut
                cell.alignment = center_alignment

    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 5,   # N°
        'B': 15,  # Matricule
        'C': 15,  # Nom
        'D': 15,  # Prénom
        'E': 15,  # Post-nom
        'F': 15,  # Téléphone
        'G': 25,  # Email
        'H': 8,   # Genre
        'I': 15,  # Date de naissance
        'J': 20,  # Département
        'K': 20,  # Ministère
        'L': 10,  # Statut
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Hauteur de la ligne d'en-tête
    ws.row_dimensions[4].height = 30

    # Figer les volets (Freeze panes) - figer l'en-tête
    ws.freeze_panes = 'A5'

    # Ajouter un filtre automatique
    ws.auto_filter.ref = f"A4:L{ws.max_row}"

    # Ajouter une ligne de total
    total_row = ws.max_row + 2
    ws.merge_cells(f'A{total_row}:B{total_row}')
    total_cell = ws[f'A{total_row}']
    total_cell.value = f"Total des membres: {len(members)}"
    total_cell.font = Font(name='Calibri', size=11, bold=True)
    total_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Créer la réponse
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="membres_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'

    buffer.close()
    return response


@login_required
def member_print_card(request, pk):
    """Imprimer la carte de membre (PDF) avec QR code"""
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, Image, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    
    member = get_object_or_404(Member, pk=pk)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="carte_membre_{member.member_number}.pdf"'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # En-tête église
    elements.append(Paragraph("ÉGLISE CONSOLATION ET PAIX DIVINE", styles['Heading2']))
    elements.append(Paragraph("Carte de Membre", styles['Heading3']))
    elements.append(Spacer(1, 0.5*cm))
    
    # Numéro de membre en grand
    id_style = ParagraphStyle(
        'IDStyle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#6366f1'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph(member.member_number, id_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Informations du membre
    data = [
        ['NOM:', member.user.last_name.upper() if member.user else ''],
        ['PRÉNOM:', member.user.first_name if member.user else ''],
        ['TÉLÉPHONE:', member.user.phone if member.user else ''],
        ['DATE DE NAISSANCE:', str(member.birth_date) if member.birth_date else ''],
        ['GENRE:', member.get_gender_display() if member.gender else ''],
        ['POSITION:', member.church_position or 'Membre'],
    ]
    
    table = Table(data, colWidths=[5*cm, 9*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#475569')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    
    # QR Code
    if member.qr_code and member.qr_code.path:
        try:
            elements.append(Paragraph("Scannez pour vérifier l'authenticité:", styles['Normal']))
            elements.append(Spacer(1, 0.3*cm))
            img = Image(member.qr_code.path, width=4*cm, height=4*cm)
            elements.append(img)
        except:
            pass
    
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph("Cette carte est strictement personnelle.", styles['Normal']))
    elements.append(Paragraph("En cas de perte, veuillez contacter l'administration.", styles['Normal']))
    
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


# ============================================================
# Familles - CRUD
# ============================================================

@login_required
def family_list(request):
    """Liste des familles"""
    families = Family.objects.all().prefetch_related('members')
    return render(request, 'dashboard/families.html', {'families': families, 'view': 'list'})


@login_required
def family_create(request):
    """Créer une famille"""
    if request.method == 'POST':
        form = FamilyForm(request.POST)
        if form.is_valid():
            family = form.save()
            messages.success(request, f'Famille {family.name} créée avec succès!')
            return redirect('family-list')
    else:
        form = FamilyForm()
    
    return render(request, 'dashboard/families.html', {'form': form, 'action': 'Créer', 'view': 'form'})


@login_required
def family_edit(request, pk):
    """Modifier une famille"""
    family = get_object_or_404(Family, pk=pk)
    
    if request.method == 'POST':
        form = FamilyForm(request.POST, instance=family)
        if form.is_valid():
            form.save()
            messages.success(request, f'Famille {family.name} modifiée avec succès!')
            return redirect('family-list')
    else:
        form = FamilyForm(instance=family)
    
    return render(request, 'dashboard/families.html', {
        'form': form,
        'family': family,
        'action': 'Modifier',
        'view': 'form'
    })


@login_required
def family_delete(request, pk):
    """Supprimer une famille"""
    family = get_object_or_404(Family, pk=pk)
    
    if request.method == 'POST':
        family.delete()
        messages.success(request, 'Famille supprimée avec succès!')
        return redirect('family-list')
    
    return render(request, 'dashboard/families.html', {'family': family, 'view': 'delete_confirm'})


# ============================================================
# Groupes de maison - CRUD
# ============================================================

@login_required
def homegroup_list(request):
    """Liste des groupes de maison"""
    homegroups = HomeGroup.objects.all().select_related('leader')
    return render(request, 'dashboard/homegroups.html', {'homegroups': homegroups, 'view': 'list'})


@login_required
def homegroup_create(request):
    """Créer un groupe de maison"""
    if request.method == 'POST':
        form = HomeGroupForm(request.POST)
        if form.is_valid():
            homegroup = form.save()
            messages.success(request, f'Groupe de maison {homegroup.name} créé avec succès!')
            return redirect('homegroup-list')
    else:
        form = HomeGroupForm()
    
    return render(request, 'dashboard/homegroups.html', {'form': form, 'action': 'Créer', 'view': 'form'})


@login_required
def homegroup_edit(request, pk):
    """Modifier un groupe de maison"""
    homegroup = get_object_or_404(HomeGroup, pk=pk)
    
    if request.method == 'POST':
        form = HomeGroupForm(request.POST, instance=homegroup)
        if form.is_valid():
            form.save()
            messages.success(request, f'Groupe de maison {homegroup.name} modifié avec succès!')
            return redirect('homegroup-list')
    else:
        form = HomeGroupForm(instance=homegroup)
    
    return render(request, 'dashboard/homegroups.html', {
        'form': form,
        'homegroup': homegroup,
        'action': 'Modifier',
        'view': 'form'
    })


@login_required
def homegroup_delete(request, pk):
    """Supprimer un groupe de maison"""
    homegroup = get_object_or_404(HomeGroup, pk=pk)
    
    if request.method == 'POST':
        homegroup.delete()
        messages.success(request, 'Groupe de maison supprimé avec succès!')
        return redirect('homegroup-list')
    
    return render(request, 'dashboard/homegroups.html', {'homegroup': homegroup, 'view': 'delete_confirm'})


# ============================================================
# Départements - CRUD
# ============================================================

@login_required
def department_list(request):
    """Liste des départements"""
    departments = Department.objects.all().select_related('head')
    return render(request, 'dashboard/departments.html', {'departments': departments, 'view': 'list'})


@login_required
def department_create(request):
    """Créer un département"""
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save()
            messages.success(request, f'Département {department.name} créé avec succès!')
            return redirect('department-list')
    else:
        form = DepartmentForm()
    
    return render(request, 'dashboard/departments.html', {'form': form, 'action': 'Créer', 'view': 'form'})


@login_required
def department_edit(request, pk):
    """Modifier un département"""
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, f'Département {department.name} modifié avec succès!')
            return redirect('department-list')
    else:
        form = DepartmentForm(instance=department)
    
    return render(request, 'dashboard/departments.html', {
        'form': form,
        'department': department,
        'action': 'Modifier',
        'view': 'form'
    })


@login_required
def department_delete(request, pk):
    """Supprimer un département"""
    department = get_object_or_404(Department, pk=pk)

    if request.method == 'POST':
        department.delete()
        messages.success(request, 'Département supprimé avec succès!')
        return redirect('department-list')

    return render(request, 'dashboard/departments.html', {'department': department, 'view': 'delete_confirm'})


# ============================================================
# Ministères - CRUD
# ============================================================

@login_required
def ministry_list(request):
    """Liste des ministères"""
    ministries = Ministry.objects.all().select_related('leader')
    return render(request, 'dashboard/ministries.html', {'ministries': ministries, 'view': 'list'})


@login_required
def ministry_create(request):
    """Créer un ministère"""
    if request.method == 'POST':
        form = MinistryForm(request.POST)
        if form.is_valid():
            ministry = form.save()
            messages.success(request, f'Ministère {ministry.name} créé avec succès!')
            return redirect('ministry-list')
    else:
        form = MinistryForm()
    
    return render(request, 'dashboard/ministries.html', {'form': form, 'action': 'Créer', 'view': 'form'})


@login_required
def ministry_edit(request, pk):
    """Modifier un ministère"""
    ministry = get_object_or_404(Ministry, pk=pk)
    
    if request.method == 'POST':
        form = MinistryForm(request.POST, instance=ministry)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ministère {ministry.name} modifié avec succès!')
            return redirect('ministry-list')
    else:
        form = MinistryForm(instance=ministry)
    
    return render(request, 'dashboard/ministries.html', {
        'form': form,
        'ministry': ministry,
        'action': 'Modifier',
        'view': 'form'
    })


@login_required
def ministry_delete(request, pk):
    """Supprimer un ministère"""
    ministry = get_object_or_404(Ministry, pk=pk)

    if request.method == 'POST':
        ministry.delete()
        messages.success(request, 'Ministère supprimé avec succès!')
        return redirect('ministry-list')

    return render(request, 'dashboard/ministries.html', {'ministry': ministry, 'view': 'delete_confirm'})


# ============================================================
# Événements - CRUD
# ============================================================

@login_required
def event_list(request):
    """Liste des événements avec pagination, recherche et filtres"""
    from datetime import date, timedelta
    
    events = Event.objects.all().order_by('-date', '-time')
    
    # Recherche
    search = request.GET.get('q', '')
    if search:
        events = events.filter(title__icontains=search)
    
    # Filtre par type
    event_type = request.GET.get('type', '')
    if event_type:
        events = events.filter(event_type=event_type)
    
    # Filtre par statut
    status_filter = request.GET.get('status', '')
    today = date.today()
    if status_filter == 'upcoming':
        events = events.filter(date__gte=today)
    elif status_filter == 'past':
        events = events.filter(date__lt=today)
    elif status_filter == 'published':
        events = events.filter(is_published=True)
    
    # Statistiques
    total_events = Event.objects.count()
    upcoming_events = Event.objects.filter(date__gte=today).count()
    this_week_start = today - timedelta(days=today.weekday())
    this_week_end = this_week_start + timedelta(days=6)
    this_week_events = Event.objects.filter(date__range=[this_week_start, this_week_end]).count()
    published_events = Event.objects.filter(is_published=True).count()
    
    # Pagination
    paginator = Paginator(events, 12)  # 12 événements par page (grid layout)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dashboard/events.html', {
        'events': page_obj,
        'event_type': event_type,
        'status_filter': status_filter,
        'search': search,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'event_types': Event.EVENT_TYPE_CHOICES,
        'total_events': total_events,
        'upcoming_events': upcoming_events,
        'this_week_events': this_week_events,
        'published_events': published_events,
    })


@login_required
def event_detail(request, pk):
    """Détail d'un événement"""
    event = get_object_or_404(Event, pk=pk)
    attendance_list = Attendance.objects.filter(event=event).select_related('member')

    # Get all active members for comparison
    total_members = Member.objects.filter(is_active=True).count()

    # Get demographic counters
    aggregate = EventAttendanceAggregate.objects.filter(event=event).first()
    demographic_total = aggregate.grand_total if aggregate else 0

    # Calculate combined attendance stats
    member_present_count = attendance_list.filter(attended=True).count()
    member_absent_count = attendance_list.filter(attended=False).count()
    total_presence = member_present_count + demographic_total
    total_expected = total_members + demographic_total
    global_rate = round((total_presence / total_expected * 100)) if total_expected > 0 else 0

    return render(request, 'dashboard/event-detail.html', {
        'event': event,
        'attendance_list': attendance_list,
        'total_attendance': attendance_list.count(),
        'present_count': member_present_count,
        'absent_count': member_absent_count,
        'not_marked_count': total_members - member_present_count - member_absent_count,
        'total_members': total_members,
        'aggregate': aggregate,
        'demographic_total': demographic_total,
        'total_presence': total_presence,
        'global_rate': global_rate,
    })


@login_required
def event_create(request):
    """Créer un événement"""
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES)
        if form.is_valid():
            event = form.save()
            messages.success(request, f'Événement "{event.title}" créé avec succès!')
            return redirect('event-detail', pk=event.pk)
        else:
            # Form has errors - return with error messages
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        form = EventForm()

    departments = Department.objects.all()
    return render(request, 'dashboard/events.html', {
        'form': form, 
        'action': 'Créer', 
        'view': 'event_form',
        'departments': departments
    })


@login_required
def event_edit(request, pk):
    """Modifier un événement"""
    event = get_object_or_404(Event, pk=pk)
    
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save()
            messages.success(request, f'Événement "{event.title}" modifié avec succès!')
            return redirect('event-detail', pk=event.pk)
    else:
        form = EventForm(instance=event)
    
    departments = Department.objects.all()
    return render(request, 'dashboard/events.html', {
        'form': form,
        'event': event,
        'action': 'Modifier',
        'view': 'event_form',
        'departments': departments
    })


@login_required
def event_delete(request, pk):
    """Supprimer un événement"""
    event = get_object_or_404(Event, pk=pk)
    
    if request.method == 'POST':
        event.delete()
        messages.success(request, 'Événement supprimé avec succès!')
        return redirect('event-list')
    
    return render(request, 'dashboard/events.html', {'event': event, 'delete_confirm': True})


# ============================================================
# Présences (Pointage)
# ============================================================

@login_required
def attendance_list(request):
    """Liste des présences"""
    attendances = Attendance.objects.all().select_related('event', 'member').order_by('-event__date')
    return render(request, 'dashboard/diaconat.html', {'attendances': attendances, 'view': 'attendance_list'})


@login_required
def attendance_event(request, event_pk):
    """Pointage pour un événement spécifique"""
    event = get_object_or_404(Event, pk=event_pk)

    # Get or create aggregate for the event (always needed for context)
    aggregate, created = EventAttendanceAggregate.objects.get_or_create(
        event=event,
        defaults={'male_adults': 0, 'female_adults': 0, 'young_men': 0, 'young_women': 0,
                 'male_children': 0, 'female_children': 0, 'elderly_men': 0, 'elderly_women': 0}
    )

    if request.method == 'POST':
        action = request.POST.get('action', '')
        member_id = request.POST.get('member_id')

        if action == 'mark_all_present':
            # Mark all active members as present
            members = Member.objects.filter(is_active=True)
            for member in members:
                attendance, created = Attendance.objects.get_or_create(
                    event=event,
                    member=member,
                    defaults={'attended': True, 'checked_in_at': timezone.now()}
                )
                if not created:
                    attendance.attended = True
                    attendance.checked_in_at = timezone.now()
                    attendance.save()
            messages.success(request, f'Tous les membres ont été marqués présents!')
            return redirect('attendance-event', event_pk=event_pk)

        elif action == 'mark_all_absent':
            # Mark all active members as absent
            members = Member.objects.filter(is_active=True)
            for member in members:
                attendance, created = Attendance.objects.get_or_create(
                    event=event,
                    member=member,
                    defaults={'attended': False, 'checked_in_at': None}
                )
                if not created:
                    attendance.attended = False
                    attendance.checked_in_at = None
                    attendance.save()
            messages.success(request, f'Tous les membres ont été marqués absents!')
            return redirect('attendance-event', event_pk=event_pk)

        elif action == 'reset':
            # Reset all attendances for this event
            Attendance.objects.filter(event=event).delete()
            messages.success(request, 'Pointage réinitialisé!')
            return redirect('attendance-event', event_pk=event_pk)

        elif member_id and action in ['present', 'absent']:
            # Mark specific member
            member = get_object_or_404(Member, pk=member_id)
            attended = action == 'present'
            attendance, created = Attendance.objects.get_or_create(
                event=event,
                member=member,
                defaults={'attended': attended, 'checked_in_at': timezone.now() if attended else None}
            )
            if not created:
                attendance.attended = attended
                attendance.checked_in_at = timezone.now() if attended else None
                attendance.save()

            # Calculate updated counts
            present_count = Attendance.objects.filter(event=event, attended=True).count()
            absent_count = Attendance.objects.filter(event=event, attended=False).count()

            # Get demographic totals for combined calculation
            aggregate = EventAttendanceAggregate.objects.filter(event=event).first()
            demographic_total = aggregate.grand_total if aggregate else 0
            total_members = Member.objects.filter(is_active=True).count()
            total_presence = present_count + demographic_total
            total_expected = total_members + demographic_total
            global_rate = round((total_presence / total_expected * 100)) if total_expected > 0 else 0

            # Check if AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'present_count': present_count,
                    'absent_count': absent_count,
                    'demographic_total': demographic_total,
                    'total_presence': total_presence,
                    'global_rate': global_rate,
                    'member_name': member.get_full_name(),
                    'action': action
                })

            messages.success(request, f'Présence enregistrée pour {member.get_full_name}!')
            return redirect('attendance-event', event_pk=event_pk)

        # Handle AttendanceForm
        form = AttendanceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Présence enregistrée avec succès!')
            return redirect('attendance-event', event_pk=event_pk)

        # Handle demographic counters form
        aggregate_form = EventAttendanceAggregateForm(request.POST, instance=aggregate)
        if aggregate_form.is_valid():
            aggregate = aggregate_form.save(commit=False)
            aggregate.updated_by = request.user
            aggregate.save()
            messages.success(request, 'Compteurs démographiques enregistrés avec succès!')
            return redirect('attendance-event', event_pk=event_pk)
    else:
        form = AttendanceForm(initial={'event': event})
        aggregate_form = EventAttendanceAggregateForm(instance=aggregate)

    # Get all active members
    members = Member.objects.filter(is_active=True).order_by('user__last_name', 'user__first_name')

    # Get existing attendances for this event
    attendances = Attendance.objects.filter(event=event).select_related('member')
    member_attendance = {a.member_id: a for a in attendances}

    # Calculate stats
    present_count = attendances.filter(attended=True).count()
    absent_count = attendances.filter(attended=False).count()
    total_members = members.count()
    not_marked_count = total_members - present_count - absent_count

    # Calculate combined attendance rate (members + demographic counters)
    demographic_total = aggregate.grand_total if aggregate else 0
    total_presence = present_count + demographic_total
    total_expected = total_members + demographic_total

    # Calculate attendance rate for progress ring (combined)
    attendance_rate = round((total_presence / total_expected * 100)) if total_expected > 0 else 0
    attendance_offset = 326.73 - (attendance_rate / 100 * 326.73)  # 326.73 is circumference of circle r=52

    # Member-only rate for reference
    member_attendance_rate = round((present_count / total_members * 100)) if total_members > 0 else 0

    return render(request, 'dashboard/attendance.html', {
        'event': event,
        'form': form,
        'aggregate_form': aggregate_form,
        'aggregate': aggregate,
        'members': members,
        'member_attendance': member_attendance,
        'present_count': present_count,
        'absent_count': absent_count,
        'not_marked_count': not_marked_count,
        'total_members': total_members,
        'attendance_rate': attendance_rate,
        'attendance_offset': attendance_offset,
        'member_attendance_rate': member_attendance_rate,
        'demographic_total': demographic_total,
        'total_presence': total_presence,
    })


# ============================================================
# Finances - CRUD
# ============================================================

@login_required
@finance_required
def finance_list(request):
    """Liste des transactions financières avec création directe"""
    from datetime import datetime

    # Handle POST request for creating transaction directly from main page
    if request.method == 'POST':
        try:
            direction = request.POST.get('direction', 'in')
            amount = request.POST.get('amount', '0')
            date_str = request.POST.get('date')
            category_id = request.POST.get('category')
            description = request.POST.get('description', '')
            currency = request.POST.get('currency', 'CDF')
            reference = request.POST.get('reference', '')
            event_id = request.POST.get('event')
            member_id = request.POST.get('member')
            transaction_type = request.POST.get('transaction_type', 'offering' if direction == 'in' else 'functioning')
            payment_method = request.POST.get('payment_method', '')
            donor_name = request.POST.get('donor_name', '')
            recipient_name = request.POST.get('recipient_name', '')

            if not amount or float(amount) <= 0:
                messages.error(request, 'Le montant doit être supérieur à 0')
                return redirect('finance-list')

            # Get category
            category = None
            if category_id:
                try:
                    category = FinancialCategory.objects.get(pk=category_id)
                except FinancialCategory.DoesNotExist:
                    pass

            # Get event
            event = None
            if event_id:
                try:
                    event = Event.objects.get(pk=event_id)
                except Event.DoesNotExist:
                    pass

            # Get member
            member = None
            if member_id:
                try:
                    member = Member.objects.get(pk=member_id)
                except Member.DoesNotExist:
                    pass

            # Auto-generate reference if not provided
            if not reference:
                prefix = 'ENT' if direction == 'in' else 'SOR'
                year = datetime.now().year
                # Get last transaction number for this year
                last_count = FinancialTransaction.objects.filter(
                    reference_number__startswith=f'{prefix}-{year}'
                ).count()
                reference = f'{prefix}-{year}-{str(last_count + 1).zfill(4)}'

            # Create transaction
            transaction = FinancialTransaction.objects.create(
                direction=direction,
                amount=float(amount),
                date=date_str or datetime.now().date(),
                category=category,
                event=event,
                member=member,
                transaction_type=transaction_type,
                payment_method=payment_method,
                description=description,
                currency=currency,
                reference_number=reference,
                document_number=reference,
                donor_name=donor_name if direction == 'in' else None,
                recipient_name=recipient_name if direction == 'out' else None,
                created_by=request.user,
                cashier=request.user
            )

            messages.success(request, f'Transaction {reference} enregistrée avec succès!')
            return redirect('finance-list')

        except Exception as e:
            messages.error(request, f'Erreur lors de l\'enregistrement: {str(e)}')
            return redirect('finance-list')

    # GET request - show list
    transactions = FinancialTransaction.objects.all().select_related('category', 'member').order_by('-date', '-created_at')

    # Calculate monthly stats
    from django.utils import timezone
    today = timezone.now()
    first_day_of_month = today.replace(day=1)

    # Calculate monthly stats by currency
    month_in_cdf = FinancialTransaction.objects.filter(
        direction='in', currency='CDF',
        date__gte=first_day_of_month
    ).aggregate(total=Sum('amount'))['total'] or 0

    month_out_cdf = FinancialTransaction.objects.filter(
        direction='out', currency='CDF',
        date__gte=first_day_of_month
    ).aggregate(total=Sum('amount'))['total'] or 0

    month_in_usd = FinancialTransaction.objects.filter(
        direction='in', currency='USD',
        date__gte=first_day_of_month
    ).aggregate(total=Sum('amount'))['total'] or 0

    month_out_usd = FinancialTransaction.objects.filter(
        direction='out', currency='USD',
        date__gte=first_day_of_month
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Calculate total balances by currency
    total_in_cdf = FinancialTransaction.objects.filter(
        direction='in', currency='CDF'
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_out_cdf = FinancialTransaction.objects.filter(
        direction='out', currency='CDF'
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_in_usd = FinancialTransaction.objects.filter(
        direction='in', currency='USD'
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_out_usd = FinancialTransaction.objects.filter(
        direction='out', currency='USD'
    ).aggregate(total=Sum('amount'))['total'] or 0

    balance_cdf = total_in_cdf - total_out_cdf
    balance_usd = total_in_usd - total_out_usd

    # Get upcoming and recent events for selection
    upcoming_events = Event.objects.filter(date__gte=timezone.now().date() - timezone.timedelta(days=7)).order_by('-date')[:20]
    members = Member.objects.filter(is_active=True).order_by('post_name', 'member_number')

    context = {
        'transactions': transactions[:100],
        'categories': FinancialCategory.objects.all(),
        'events': upcoming_events,
        'members': members,
        'transaction_types': FinancialTransaction.TRANSACTION_TYPE_CHOICES,
        'payment_methods': [
            ('cash', 'Espèces'),
            ('bank_transfer', 'Virement bancaire'),
            ('mobile_money', 'Mobile Money'),
            ('check', 'Chèque'),
            ('card', 'Carte bancaire'),
        ],
        # CDF totals
        'total_in_cdf': total_in_cdf,
        'total_out_cdf': total_out_cdf,
        'balance_cdf': balance_cdf,
        'month_in_cdf': month_in_cdf,
        'month_out_cdf': month_out_cdf,
        # USD totals
        'total_in_usd': total_in_usd,
        'total_out_usd': total_out_usd,
        'balance_usd': balance_usd,
        'month_in_usd': month_in_usd,
        'month_out_usd': month_out_usd,
        # Combined totals for backwards compatibility
        'total_in': total_in_cdf + total_in_usd,
        'total_out': total_out_cdf + total_out_usd,
        'month_in': month_in_cdf + month_in_usd,
        'month_out': month_out_cdf + month_out_usd,
    }
    return render(request, 'dashboard/finances.html', context)


@login_required
@finance_required
def transaction_create(request):
    """Créer une transaction"""
    if request.method == 'POST':
        form = FinancialTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save()
            messages.success(request, f'Transaction {transaction.description} enregistrée avec succès!')
            return redirect('finance-list')
    else:
        form = FinancialTransactionForm()
    
    return render(request, 'dashboard/finances.html', {'form': form, 'action': 'Créer', 'view': 'transaction_form'})


@login_required
@finance_required
def transaction_edit(request, pk):
    """Modifier une transaction"""
    transaction = get_object_or_404(FinancialTransaction, pk=pk)

    if request.method == 'POST':
        # Handle manual form submission
        try:
            transaction.direction = request.POST.get('direction', transaction.direction)
            transaction.amount = request.POST.get('amount', transaction.amount)
            transaction.currency = request.POST.get('currency', transaction.currency)
            transaction.date = request.POST.get('date', transaction.date)
            transaction.transaction_type = request.POST.get('transaction_type', transaction.transaction_type)
            transaction.payment_method = request.POST.get('payment_method', transaction.payment_method)
            transaction.description = request.POST.get('description', transaction.description)
            transaction.reference_number = request.POST.get('reference_number', transaction.reference_number)
            transaction.donor_name = request.POST.get('donor_name', '')
            transaction.recipient_name = request.POST.get('recipient_name', '')

            # Foreign keys
            category_id = request.POST.get('category')
            if category_id:
                transaction.category = FinancialCategory.objects.filter(pk=category_id).first()

            event_id = request.POST.get('event')
            if event_id:
                transaction.event = Event.objects.filter(pk=event_id).first()
            else:
                transaction.event = None

            member_id = request.POST.get('member')
            if member_id:
                transaction.member = Member.objects.filter(pk=member_id).first()
            else:
                transaction.member = None

            transaction.save()
            messages.success(request, f'Transaction {transaction.reference_number} modifiée avec succès!')
            return redirect('finance-list')

        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')

    # Get context data for form
    upcoming_events = Event.objects.filter(date__gte=timezone.now().date() - timezone.timedelta(days=7)).order_by('-date')[:20]
    members = Member.objects.filter(is_active=True).order_by('post_name', 'member_number')

    return render(request, 'dashboard/finances.html', {
        'transaction': transaction,
        'categories': FinancialCategory.objects.all(),
        'events': upcoming_events,
        'members': members,
        'transaction_types': FinancialTransaction.TRANSACTION_TYPE_CHOICES,
        'payment_methods': [
            ('cash', 'Espèces'),
            ('bank_transfer', 'Virement bancaire'),
            ('mobile_money', 'Mobile Money'),
            ('check', 'Chèque'),
            ('card', 'Carte bancaire'),
        ],
        'action': 'Modifier',
        'view': 'transaction_form'
    })


@login_required
@finance_required
def transaction_delete(request, pk):
    """Supprimer une transaction"""
    transaction = get_object_or_404(FinancialTransaction, pk=pk)
    
    if request.method == 'POST':
        transaction.delete()
        messages.success(request, 'Transaction supprimée avec succès!')
        return redirect('finance-list')
    
    return render(request, 'dashboard/finances.html', {'transaction': transaction, 'delete_confirm': True})


@login_required
@finance_required
def transaction_detail(request, pk):
    """Détail d'une transaction"""
    transaction = get_object_or_404(FinancialTransaction, pk=pk)
    return render(request, 'dashboard/finances.html', {
        'transaction': transaction,
        'view': 'transaction_detail'
    })


@login_required
@finance_required
def category_list(request):
    """Liste des catégories financières"""
    categories = FinancialCategory.objects.all()
    return render(request, 'dashboard/finances.html', {'categories': categories, 'view': 'category_list'})


@login_required
@finance_required
def category_create(request):
    """Créer une catégorie"""
    if request.method == 'POST':
        form = FinancialCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Catégorie "{category.name}" créée avec succès!')
            return redirect('finance-list')
    else:
        form = FinancialCategoryForm()

    return render(request, 'dashboard/finances.html', {'form': form, 'action': 'Créer', 'view': 'category_form'})


@login_required
@finance_required
def category_edit(request, pk):
    """Modifier une catégorie"""
    category = get_object_or_404(FinancialCategory, pk=pk)
    
    if request.method == 'POST':
        form = FinancialCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, f'Catégorie {category.name} modifiée avec succès!')
            return redirect('category-list')
    else:
        form = FinancialCategoryForm(instance=category)
    
    return render(request, 'dashboard/finances.html', {
        'form': form,
        'category': category,
        'action': 'Modifier',
        'view': 'category_form'
    })


@login_required
def category_delete(request, pk):
    """Supprimer une catégorie"""
    category = get_object_or_404(FinancialCategory, pk=pk)
    
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Catégorie supprimée avec succès!')
        return redirect('category-list')
    
    return render(request, 'dashboard/finances.html', {'category': category, 'delete_confirm': True, 'view': 'category_list'})


# ============================================================
# Rapports
# ============================================================

@login_required
def reports(request):
    """Rapports et statistiques complets avec filtres de période"""
    # Récupérer les paramètres de période
    period = request.GET.get('period', 'month')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Calculer les dates par défaut selon la période
    today = timezone.now().date()
    if period == 'week':
        default_from = today - timedelta(days=7)
    elif period == 'month':
        default_from = today - timedelta(days=30)
    elif period == 'quarter':
        default_from = today - timedelta(days=90)
    elif period == 'year':
        default_from = today - timedelta(days=365)
    else:
        default_from = today - timedelta(days=30)

    # Utiliser les dates fournies ou les défauts
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        except:
            from_date = default_from
    else:
        from_date = default_from

    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        except:
            to_date = today
    else:
        to_date = today

    # Statistiques membres
    members_total = Member.objects.count()
    members_new = Member.objects.filter(created_at__date__range=[from_date, to_date]).count()
    members_active = Member.objects.filter(is_active=True).count()
    members_by_gender = {
        'male': Member.objects.filter(gender='male').count(),
        'female': Member.objects.filter(gender='female').count(),
    }
    members_by_status = {
        'active': Member.objects.filter(is_active=True, inactive_reason__isnull=True).count(),
        'inactive': Member.objects.filter(is_active=False).count(),
        'visitor': Member.objects.filter(is_active=True, inactive_reason='visitor').count(),
    }

    # Statistiques présences
    attendance_records = Attendance.objects.filter(event__date__range=[from_date, to_date])
    attendance_total = attendance_records.count()
    attendance_present = attendance_records.filter(attended=True).count()
    attendance_rate = round((attendance_present / attendance_total * 100), 1) if attendance_total > 0 else 0

    # Statistiques finances
    transactions_period = FinancialTransaction.objects.filter(date__range=[from_date, to_date])
    finances_in = transactions_period.filter(direction='in').aggregate(total=Sum('amount'))['total'] or 0
    finances_out = transactions_period.filter(direction='out').aggregate(total=Sum('amount'))['total'] or 0
    finances_total_in = FinancialTransaction.objects.filter(direction='in').aggregate(total=Sum('amount'))['total'] or 0
    finances_total_out = FinancialTransaction.objects.filter(direction='out').aggregate(total=Sum('amount'))['total'] or 0

    # Statistiques événements
    events_total = Event.objects.filter(date__range=[from_date, to_date]).count()
    events_all = Event.objects.count()
    upcoming_events = Event.objects.filter(date__gte=today).count()

    # Autres statistiques
    families_total = Family.objects.count()
    departments_total = Department.objects.count()
    ministries_total = Ministry.objects.count()
    homegroups_total = HomeGroup.objects.count()
    marriages_total = MarriageRecord.objects.count()
    baptisms_total = BaptismEvent.objects.count()
    trainings_total = TrainingEvent.objects.count()
    documents_total = Document.objects.count()
    announcements_total = Announcement.objects.count()

    # Statistiques par mois pour graphiques (6 derniers mois)
    monthly_stats = []
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=i*30)
        month_start = month_date.replace(day=1)
        month_members = Member.objects.filter(created_at__year=month_date.year, created_at__month=month_date.month).count()
        month_events = Event.objects.filter(date__year=month_date.year, date__month=month_date.month).count()
        month_finances = FinancialTransaction.objects.filter(
            date__year=month_date.year, date__month=month_date.month, direction='in'
        ).aggregate(total=Sum('amount'))['total'] or 0
        monthly_stats.append({
            'month': month_date.strftime('%b'),
            'members': month_members,
            'events': month_events,
            'finances': month_finances
        })

    context = {
        'period': period,
        'date_from': from_date,
        'date_to': to_date,
        'stats': {
            'members_total': members_total,
            'members_new': members_new,
            'members_active': members_active,
            'members_by_gender': members_by_gender,
            'members_by_status': members_by_status,
            'attendance_total': attendance_total,
            'attendance_present': attendance_present,
            'attendance_rate': attendance_rate,
            'finances_in': finances_in,
            'finances_out': finances_out,
            'finances_balance': finances_in - finances_out,
            'finances_total_in': finances_total_in,
            'finances_total_out': finances_total_out,
            'events_total': events_total,
            'events_all': events_all,
            'upcoming_events': upcoming_events,
            'families_total': families_total,
            'departments_total': departments_total,
            'ministries_total': ministries_total,
            'homegroups_total': homegroups_total,
            'marriages_total': marriages_total,
            'baptisms_total': baptisms_total,
            'trainings_total': trainings_total,
            'documents_total': documents_total,
            'announcements_total': announcements_total,
        },
        'monthly_stats': monthly_stats,
    }
    return render(request, 'dashboard/reports.html', context)


@login_required
def report_members_detail(request):
    """Rapport détaillé des membres"""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    members = Member.objects.all().select_related('family', 'department', 'home_group')

    if date_from and date_to:
        members = members.filter(created_at__date__range=[date_from, date_to])

    context = {
        'members': members,
        'date_from': date_from,
        'date_to': date_to,
        'total': members.count(),
        'by_gender': {
            'male': members.filter(gender='male').count(),
            'female': members.filter(gender='female').count(),
        },
        'by_status': {
            'active': members.filter(is_active=True, inactive_reason__isnull=True).count(),
            'inactive': members.filter(is_active=False).count(),
        }
    }
    return render(request, 'dashboard/report_members_detail.html', context)


@login_required
def report_finances_detail(request):
    """Rapport détaillé des finances"""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    transactions = FinancialTransaction.objects.all().select_related('category', 'member')

    if date_from and date_to:
        transactions = transactions.filter(date__range=[date_from, date_to])

    context = {
        'transactions': transactions,
        'date_from': date_from,
        'date_to': date_to,
        'total_in': transactions.filter(direction='in').aggregate(total=Sum('amount'))['total'] or 0,
        'total_out': transactions.filter(direction='out').aggregate(total=Sum('amount'))['total'] or 0,
    }
    return render(request, 'dashboard/report_finances_detail.html', context)


@login_required
def report_activities_detail(request):
    """Rapport détaillé des activités"""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    events = Event.objects.all().prefetch_related('attendances')

    if date_from and date_to:
        events = events.filter(date__range=[date_from, date_to])

    context = {
        'events': events,
        'date_from': date_from,
        'date_to': date_to,
        'total_events': events.count(),
    }
    return render(request, 'dashboard/report_activities_detail.html', context)


@login_required
def report_attendance_detail(request):
    """Rapport détaillé des présences"""
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    attendances = Attendance.objects.all().select_related('event', 'member')

    if date_from and date_to:
        attendances = attendances.filter(event__date__range=[date_from, date_to])

    total = attendances.count()
    present = attendances.filter(attended=True).count()
    rate = round((present / total * 100), 1) if total > 0 else 0

    context = {
        'attendances': attendances,
        'date_from': date_from,
        'date_to': date_to,
        'total': total,
        'present': present,
        'rate': rate,
    }
    return render(request, 'dashboard/report_attendance_detail.html', context)


@login_required
def report_sacraments_detail(request):
    """Rapport détaillé des sacrements (mariages et baptêmes)"""
    marriages = MarriageRecord.objects.all().select_related('groom', 'bride')
    baptisms = BaptismEvent.objects.all().prefetch_related('candidates')

    context = {
        'marriages': marriages,
        'baptisms': baptisms,
        'marriages_count': marriages.count(),
        'baptisms_count': baptisms.count(),
    }
    return render(request, 'dashboard/report_sacraments_detail.html', context)


@login_required
def export_reports_excel(request):
    """Export global des rapports en fichier Excel XLSX formaté"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    # Créer le workbook
    wb = Workbook()

    # Style commun
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=14)

    # ===== FEUILLE 1: RAPPORT GLOBAL =====
    ws_summary = wb.active
    ws_summary.title = "Rapport Global"

    # Titre
    ws_summary['A1'] = "RAPPORT GLOBAL DE L'ÉGLISE"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'].alignment = Alignment(horizontal="center")

    today = timezone.now().date()
    ws_summary['A2'] = f"Généré le {today.strftime('%d/%m/%Y')}"
    ws_summary['A2'].alignment = Alignment(horizontal="center")
    ws_summary.merge_cells('A2:D2')

    # Statistiques générales
    row = 4
    ws_summary[f'A{row}'] = "CATÉGORIE"
    ws_summary[f'B{row}'] = "TOTAL"
    ws_summary[f'C{row}'] = "ACTIFS"
    ws_summary[f'D{row}'] = "NOTES"

    for col in ['A', 'B', 'C', 'D']:
        cell = ws_summary[f'{col}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Données
    data = [
        ["Membres", Member.objects.count(), Member.objects.filter(is_active=True).count(), "Total inscrits"],
        ["Familles", Family.objects.count(), "-", "Nombre de familles"],
        ["Groupes de maison", HomeGroup.objects.count(), "-", "Cellules"],
        ["Départements", Department.objects.count(), "-", "Ministères"],
        ["Événements", Event.objects.count(), Event.objects.filter(date__gte=today).count(), "À venir"],
        ["Mariages", MarriageRecord.objects.count(), "-", "Cérémonies"],
        ["Baptêmes", BaptismEvent.objects.count(), "-", "Événements"],
        ["Formations", TrainingEvent.objects.count(), "-", "Sessions"],
        ["Documents", Document.objects.count(), "-", "Fichiers"],
    ]

    for i, row_data in enumerate(data, start=5):
        for j, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Finances
    ws_summary['A15'] = "FINANCES"
    ws_summary['A15'].font = title_font
    ws_summary.merge_cells('A15:D15')

    row = 16
    ws_summary[f'A{row}'] = "TYPE"
    ws_summary[f'B{row}'] = "MONTANT ($)"
    ws_summary[f'C{row}'] = "DÉTAILS"

    for col in ['A', 'B', 'C']:
        cell = ws_summary[f'{col}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    finances_data = [
        ["Entrées totales", FinancialTransaction.objects.filter(direction='in').aggregate(total=Sum('amount'))['total'] or 0, "Dons et contributions"],
        ["Sorties totales", FinancialTransaction.objects.filter(direction='out').aggregate(total=Sum('amount'))['total'] or 0, "Dépenses"],
        ["Solde", (FinancialTransaction.objects.filter(direction='in').aggregate(total=Sum('amount'))['total'] or 0) -
                  (FinancialTransaction.objects.filter(direction='out').aggregate(total=Sum('amount'))['total'] or 0), "Balance"],
    ]

    for i, row_data in enumerate(finances_data, start=17):
        for j, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if j == 2 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    # Ajuster les largeurs
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 30

    # ===== FEUILLE 2: LISTE DES MEMBRES =====
    ws_members = wb.create_sheet("Membres")

    headers = ["N° Membre", "Nom", "Prénom", "Genre", "Téléphone", "Email", "Famille", "Département", "Statut"]
    for col, header in enumerate(headers, 1):
        cell = ws_members.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    members = Member.objects.all().select_related('family', 'department', 'user')
    for i, member in enumerate(members, start=2):
        ws_members.cell(row=i, column=1, value=member.member_number).border = thin_border
        ws_members.cell(row=i, column=2, value=member.user.last_name if member.user else "").border = thin_border
        ws_members.cell(row=i, column=3, value=member.user.first_name if member.user else "").border = thin_border
        ws_members.cell(row=i, column=4, value=member.get_gender_display() if member.gender else "").border = thin_border
        ws_members.cell(row=i, column=5, value=member.user.phone if member.user else "").border = thin_border
        ws_members.cell(row=i, column=6, value=member.user.email if member.user else "").border = thin_border
        ws_members.cell(row=i, column=7, value=member.family.name if member.family else "").border = thin_border
        ws_members.cell(row=i, column=8, value=member.department.name if member.department else "").border = thin_border
        ws_members.cell(row=i, column=9, value="Actif" if member.is_active else "Inactif").border = thin_border

    for col in range(1, 10):
        ws_members.column_dimensions[get_column_letter(col)].width = 18

    # ===== FEUILLE 3: FINANCES =====
    ws_finances = wb.create_sheet("Finances")

    headers = ["Date", "Description", "Catégorie", "Membre", "Type", "Montant ($)", "Méthode"]
    for col, header in enumerate(headers, 1):
        cell = ws_finances.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    transactions = FinancialTransaction.objects.all().select_related('category', 'member', 'member__user')
    for i, trans in enumerate(transactions, start=2):
        ws_finances.cell(row=i, column=1, value=trans.date.strftime('%d/%m/%Y')).border = thin_border
        ws_finances.cell(row=i, column=2, value=trans.description).border = thin_border
        ws_finances.cell(row=i, column=3, value=trans.category.name if trans.category else "").border = thin_border
        ws_finances.cell(row=i, column=4, value=trans.member.get_full_name() if trans.member else "").border = thin_border
        ws_finances.cell(row=i, column=5, value="Entrée" if trans.direction == 'in' else "Sortie").border = thin_border
        cell_montant = ws_finances.cell(row=i, column=6, value=float(trans.amount))
        cell_montant.border = thin_border
        cell_montant.number_format = '#,##0.00'
        ws_finances.cell(row=i, column=7, value=trans.payment_method or "").border = thin_border

    for col in range(1, 8):
        ws_finances.column_dimensions[get_column_letter(col)].width = 18

    # ===== FEUILLE 4: ÉVÉNEMENTS =====
    ws_events = wb.create_sheet("Événements")

    headers = ["Date", "Titre", "Type", "Lieu", "Heure", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws_events.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    events = Event.objects.all()
    for i, event in enumerate(events, start=2):
        ws_events.cell(row=i, column=1, value=event.date.strftime('%d/%m/%Y')).border = thin_border
        ws_events.cell(row=i, column=2, value=event.title).border = thin_border
        ws_events.cell(row=i, column=3, value=event.event_type).border = thin_border
        ws_events.cell(row=i, column=4, value=event.location).border = thin_border
        ws_events.cell(row=i, column=5, value=str(event.time)).border = thin_border
        ws_events.cell(row=i, column=6, value=event.description[:100] if event.description else "").border = thin_border

    for col in range(1, 7):
        ws_events.column_dimensions[get_column_letter(col)].width = 20

    # Générer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=rapport_global_{today.strftime("%Y%m%d")}.xlsx'

    wb.save(response)
    return response


# ============================================================
# Annonces - CRUD
# ============================================================

@login_required
def announcement_list(request):
    """Liste des annonces"""
    if request.method == 'POST':
        # Handle create from main page
        try:
            title = request.POST.get('title')
            content = request.POST.get('content')
            is_active = request.POST.get('is_active') == 'on'

            if not title or not content:
                messages.error(request, 'Le titre et le contenu sont requis')
                return redirect('announcement-list')

            announcement = Announcement.objects.create(
                title=title,
                content=content,
                author=request.user,
                is_active=is_active
            )

            # Handle image upload
            if request.FILES.get('image'):
                announcement.image = request.FILES['image']
                announcement.save()

            messages.success(request, f'Annonce "{title}" créée avec succès!')
            return redirect('announcement-list')

        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {str(e)}')
            return redirect('announcement-list')

    # GET request
    announcements = Announcement.objects.all().order_by('-created_at')
    total_count = announcements.count()
    published_count = announcements.filter(is_active=True).count()
    draft_count = announcements.filter(is_active=False).count()

    return render(request, 'dashboard/announcements.html', {
        'announcements': announcements,
        'total_count': total_count,
        'published_count': published_count,
        'draft_count': draft_count,
    })


@login_required
def announcement_detail(request, pk):
    """Détail d'une annonce"""
    announcement = get_object_or_404(Announcement, pk=pk)
    return render(request, 'dashboard/announcements.html', {
        'announcement': announcement,
        'view': 'detail'
    })


@login_required
def announcement_create(request):
    """Créer une annonce - formulaire dédié"""
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            content = request.POST.get('content')
            is_active = request.POST.get('is_active') == 'on'

            if not title or not content:
                messages.error(request, 'Le titre et le contenu sont requis')
                return render(request, 'dashboard/announcements.html', {
                    'view': 'form',
                    'action': 'Créer',
                    'form_data': request.POST
                })

            announcement = Announcement.objects.create(
                title=title,
                content=content,
                author=request.user,
                is_active=is_active
            )

            # Handle image upload
            if request.FILES.get('image'):
                announcement.image = request.FILES['image']
                announcement.save()

            messages.success(request, f'Annonce "{title}" créée avec succès!')
            return redirect('announcement-list')

        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {str(e)}')

    return render(request, 'dashboard/announcements.html', {
        'view': 'form',
        'action': 'Créer'
    })


@login_required
def announcement_edit(request, pk):
    """Modifier une annonce"""
    announcement = get_object_or_404(Announcement, pk=pk)

    if request.method == 'POST':
        try:
            announcement.title = request.POST.get('title', announcement.title)
            announcement.content = request.POST.get('content', announcement.content)
            announcement.is_active = request.POST.get('is_active') == 'on'

            # Handle image upload
            if request.FILES.get('image'):
                announcement.image = request.FILES['image']

            announcement.save()

            messages.success(request, f'Annonce "{announcement.title}" modifiée avec succès!')
            return redirect('announcement-list')

        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')

    return render(request, 'dashboard/announcements.html', {
        'announcement': announcement,
        'action': 'Modifier',
        'view': 'form'
    })


@login_required
def announcement_delete(request, pk):
    """Supprimer une annonce"""
    announcement = get_object_or_404(Announcement, pk=pk)

    if request.method == 'POST':
        announcement.delete()
        messages.success(request, 'Annonce supprimée avec succès!')
        return redirect('announcement-list')

    return render(request, 'dashboard/announcements.html', {
        'announcement': announcement,
        'view': 'delete_confirm'
    })


# ============================================================
# Diaconat (Pointage et Logistique)
# ============================================================

@login_required
def diaconat(request):
    """Diaconat - Pointage et logistique"""
    today = timezone.now().date()

    # Get all events for the dropdown (upcoming and past 30 days)
    recent_date = today - timezone.timedelta(days=30)
    events = Event.objects.filter(date__gte=recent_date).order_by('-date')

    # Get selected event from query param
    selected_event_id = request.GET.get('event')
    selected_event = None
    members = []
    member_attendance = {}
    present_count = 0
    absent_count = 0
    total_members = 0
    not_marked_count = 0
    attendance_rate = 0
    attendance_offset = 326.73

    # Handle POST actions for attendance
    if request.method == 'POST':
        action = request.POST.get('action', '')
        selected_event_id = request.GET.get('event') or request.POST.get('event_id')

        if selected_event_id:
            try:
                selected_event = Event.objects.get(pk=selected_event_id)

                if action == 'mark_all_present':
                    members = Member.objects.filter(is_active=True)
                    for member in members:
                        attendance, created = Attendance.objects.get_or_create(
                            event=selected_event,
                            member=member,
                            defaults={'attended': True, 'checked_in_at': timezone.now()}
                        )
                        if not created:
                            attendance.attended = True
                            attendance.checked_in_at = timezone.now()
                            attendance.save()
                    messages.success(request, 'Tous les membres ont été marqués présents!')
                    return redirect(f'{request.path}?event={selected_event_id}')

                elif action == 'mark_all_absent':
                    members = Member.objects.filter(is_active=True)
                    for member in members:
                        attendance, created = Attendance.objects.get_or_create(
                            event=selected_event,
                            member=member,
                            defaults={'attended': False, 'checked_in_at': None}
                        )
                        if not created:
                            attendance.attended = False
                            attendance.checked_in_at = None
                            attendance.save()
                    messages.success(request, 'Tous les membres ont été marqués absents!')
                    return redirect(f'{request.path}?event={selected_event_id}')

                elif action == 'reset':
                    Attendance.objects.filter(event=selected_event).delete()
                    messages.success(request, 'Pointage réinitialisé!')
                    return redirect(f'{request.path}?event={selected_event_id}')

                elif action in ['present', 'absent']:
                    member_id = request.POST.get('member_id')
                    if member_id:
                        member = get_object_or_404(Member, pk=member_id)
                        attended = action == 'present'
                        attendance, created = Attendance.objects.get_or_create(
                            event=selected_event,
                            member=member,
                            defaults={'attended': attended, 'checked_in_at': timezone.now() if attended else None}
                        )
                        if not created:
                            attendance.attended = attended
                            attendance.checked_in_at = timezone.now() if attended else None
                            attendance.save()

                        # Return JSON response for AJAX
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            present_count = Attendance.objects.filter(event=selected_event, attended=True).count()
                            absent_count = Attendance.objects.filter(event=selected_event, attended=False).count()
                            return JsonResponse({
                                'success': True,
                                'present_count': present_count,
                                'absent_count': absent_count,
                                'member_name': member.get_full_name(),
                                'action': action
                            })

                        messages.success(request, f'Présence enregistrée pour {member.get_full_name()}!')
                        return redirect(f'{request.path}?event={selected_event_id}')

            except Event.DoesNotExist:
                pass

    # Load event data for display
    if selected_event_id:
        try:
            selected_event = Event.objects.get(pk=selected_event_id)

            # Get all active members
            members = Member.objects.filter(is_active=True).order_by('user__last_name', 'user__first_name')
            total_members = members.count()

            # Get existing attendances for this event
            attendances = Attendance.objects.filter(event=selected_event).select_related('member')
            member_attendance = {a.member_id: a for a in attendances}

            # Calculate stats
            present_count = attendances.filter(attended=True).count()
            absent_count = attendances.filter(attended=False).count()
            not_marked_count = total_members - present_count - absent_count

            # Calculate attendance rate
            total_marked = present_count + absent_count
            attendance_rate = round((present_count / total_marked * 100)) if total_marked > 0 else 0

            # Calculate progress ring offset
            circumference = 326.73
            attendance_offset = circumference - (attendance_rate / 100) * circumference

        except Event.DoesNotExist:
            pass

    # Récupérer les catégories et états dynamiques depuis la BD
    logistics_categories_db = LogisticsCategory.objects.filter(is_active=True).values_list('code', 'name')
    logistics_conditions_db = LogisticsCondition.objects.filter(is_active=True).values_list('code', 'name')

    # Catégories par défaut + celles de la BD
    default_categories = [
        ('furniture', 'Mobilier'),
        ('equipment', 'Équipement'),
        ('consumable', 'Consommable'),
        ('other', 'Autre'),
    ]
    logistics_categories = list(logistics_categories_db) or default_categories

    # États par défaut + ceux de la BD
    default_conditions = LogisticsItem.CONDITION_CHOICES
    condition_choices = list(logistics_conditions_db) or default_conditions

    context = {
        'events': events,
        'selected_event': selected_event,
        'members': members,
        'member_attendance': member_attendance,
        'present_count': present_count,
        'absent_count': absent_count,
        'total_members': total_members,
        'not_marked_count': not_marked_count,
        'attendance_rate': attendance_rate,
        'attendance_offset': attendance_offset,
        'logistics_items': LogisticsItem.objects.filter(is_active=True).order_by('-created_at')[:50],
        'logistics_categories': logistics_categories,
        'condition_choices': condition_choices,
        'currency_choices': LogisticsItem.CURRENCY_CHOICES,
    }
    return render(request, 'dashboard/diaconat.html', context)


@login_required
def diaconat_attendance(request):
    """Pointage pour le diaconat"""
    today = timezone.now().date()
    events = Event.objects.filter(date=today).order_by('time')
    return render(request, 'dashboard/diaconat.html', {'events': events, 'view': 'diaconat_attendance'})


# ============================================================
# Logistique - CRUD
# ============================================================

@login_required
def logistics_list(request):
    """Liste des éléments logistiques"""
    items = LogisticsItem.objects.filter(is_active=True).order_by('-created_at')[:50]
    return render(request, 'dashboard/diaconat.html', {
        'logistics_items': items,
        'logistics_categories': [
            ('furniture', 'Mobilier'),
            ('equipment', 'Équipement'),
            ('consumable', 'Consommable'),
            ('other', 'Autre'),
        ],
        'condition_choices': LogisticsItem.CONDITION_CHOICES,
        'currency_choices': LogisticsItem.CURRENCY_CHOICES,
    })


@login_required
def logistics_create(request):
    """Créer un élément logistique avec tous les champs du modèle"""
    if request.method == 'POST':
        try:
            item = LogisticsItem.objects.create(
                name=request.POST.get('name'),
                category=request.POST.get('category', ''),
                asset_tag=request.POST.get('asset_tag', '') or None,
                quantity=int(request.POST.get('quantity', 1)),
                unit=request.POST.get('unit', '') or None,
                condition=request.POST.get('condition', 'good'),
                location=request.POST.get('location', '') or None,
                acquired_date=request.POST.get('acquired_date') or None,
                unit_price=request.POST.get('unit_price') or None,
                purchase_currency=request.POST.get('purchase_currency', 'CDF'),
                supplier=request.POST.get('supplier', '') or None,
                notes=request.POST.get('notes', '') or None,
                is_active=request.POST.get('is_active') == 'on'
            )
            messages.success(request, f'Article "{item.name}" créé avec succès!')
            return redirect('diaconat')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {str(e)}')

    return redirect('diaconat')


@login_required
def logistics_edit(request, pk):
    """Modifier un élément logistique"""
    item = get_object_or_404(LogisticsItem, pk=pk)

    if request.method == 'POST':
        try:
            item.name = request.POST.get('name', item.name)
            item.category = request.POST.get('category', item.category)
            item.asset_tag = request.POST.get('asset_tag') or None
            item.quantity = int(request.POST.get('quantity', item.quantity))
            item.unit = request.POST.get('unit') or None
            item.condition = request.POST.get('condition', item.condition)
            item.location = request.POST.get('location') or None
            item.acquired_date = request.POST.get('acquired_date') or None
            item.unit_price = request.POST.get('unit_price') or None
            item.purchase_currency = request.POST.get('purchase_currency', item.purchase_currency)
            item.supplier = request.POST.get('supplier') or None
            item.notes = request.POST.get('notes') or None
            item.is_active = request.POST.get('is_active') == 'on'
            item.save()
            messages.success(request, f'Article "{item.name}" modifié avec succès!')
            return redirect('diaconat')
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')

    return redirect('diaconat')


@login_required
@require_http_methods(['POST'])
def logistics_create_category_ajax(request):
    """Créer une nouvelle catégorie via AJAX"""
    name = request.POST.get('name', '').strip()
    if name:
        code = name.lower().replace(' ', '_')
        category, created = LogisticsCategory.objects.get_or_create(
            code=code,
            defaults={'name': name}
        )
        return JsonResponse({
            'success': True,
            'code': category.code,
            'name': category.name,
            'created': created
        })
    return JsonResponse({'success': False, 'error': 'Nom requis'})


@login_required
@require_http_methods(['POST'])
def logistics_create_condition_ajax(request):
    """Créer un nouvel état via AJAX"""
    name = request.POST.get('name', '').strip()
    if name:
        code = name.lower().replace(' ', '_')
        condition, created = LogisticsCondition.objects.get_or_create(
            code=code,
            defaults={'name': name}
        )
        return JsonResponse({
            'success': True,
            'code': condition.code,
            'name': condition.name,
            'created': created
        })
    return JsonResponse({'success': False, 'error': 'Nom requis'})


@login_required
def logistics_delete(request, pk):
    """Supprimer un élément logistique"""
    item = get_object_or_404(LogisticsItem, pk=pk)

    if request.method == 'POST':
        item.delete()
        messages.success(request, f'Article "{item.name}" supprimé avec succès!')
        return redirect('diaconat')

    return redirect('diaconat')


@login_required
def logistics_detail(request, pk):
    """Détail d'un élément logistique"""
    item = get_object_or_404(LogisticsItem, pk=pk)
    return render(request, 'dashboard/diaconat.html', {
        'logistics_item': item,
        'view': 'logistics_detail'
    })


# ============================================================
# AJAX Endpoints pour création dynamique (Famille, Département, Ministère)
# ============================================================

@login_required
@require_http_methods(['POST'])
def ajax_create_family(request):
    """Créer une nouvelle famille via AJAX"""
    name = request.POST.get('name', '').strip()
    if name:
        family, created = Family.objects.get_or_create(
            name=name,
            defaults={'address': '', 'phone': ''}
        )
        return JsonResponse({
            'success': True,
            'id': family.id,
            'name': family.name,
            'created': created
        })
    return JsonResponse({'success': False, 'error': 'Nom de famille requis'})


@login_required
@require_http_methods(['POST'])
def ajax_create_department(request):
    """Créer un nouveau département via AJAX"""
    name = request.POST.get('name', '').strip()
    if name:
        department, created = Department.objects.get_or_create(
            name=name,
            defaults={'description': ''}
        )
        return JsonResponse({
            'success': True,
            'id': department.id,
            'name': department.name,
            'created': created
        })
    return JsonResponse({'success': False, 'error': 'Nom du département requis'})


@login_required
@require_http_methods(['POST'])
def ajax_create_ministry(request):
    """Créer un nouveau ministère via AJAX"""
    name = request.POST.get('name', '').strip()
    if name:
        ministry, created = Ministry.objects.get_or_create(
            name=name,
            defaults={'description': ''}
        )
        return JsonResponse({
            'success': True,
            'id': ministry.id,
            'name': ministry.name,
            'created': created
        })
    return JsonResponse({'success': False, 'error': 'Nom du ministère requis'})


# ============================================================
# Évangélisation - CRUD
# ============================================================

@login_required
def evangelisation_list(request):
    """Liste des activités d'évangélisation"""
    activities = EvangelismActivity.objects.all().order_by('-date', '-time')
    activity_choices = EvangelismActivity.ACTIVITY_TYPE_CHOICES
    
    # Compter les membres non baptisés pour alerte
    unbaptized_count = Member.objects.filter(is_active=True, baptism_date__isnull=True).count()
    
    return render(request, 'dashboard/evangelisation.html', {
        'activities': activities,
        'activity_choices': activity_choices,
        'unbaptized_count': unbaptized_count
    })


@login_required
def evangelisation_create(request):
    """Créer une activité d'évangélisation"""
    if request.method == 'POST':
        try:
            activity = EvangelismActivity.objects.create(
                title=request.POST.get('title'),
                activity_type=request.POST.get('activity_type', 'field'),
                custom_activity_type=request.POST.get('custom_activity_type'),
                date=request.POST.get('date'),
                time=request.POST.get('time'),
                location=request.POST.get('location'),
                moderator=request.POST.get('moderator'),
                created_by=request.user
            )
            messages.success(request, f'Activité {activity.title} créée avec succès!')
            return redirect('evangelisation-list')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {str(e)}')
    
    activity_choices = EvangelismActivity.ACTIVITY_TYPE_CHOICES
    return render(request, 'dashboard/evangelisation.html', {
        'action': 'Créer',
        'activity_choices': activity_choices
    })


@login_required
def evangelisation_edit(request, pk):
    """Modifier une activité d'évangélisation"""
    activity = get_object_or_404(EvangelismActivity, pk=pk)
    
    if request.method == 'POST':
        try:
            activity.title = request.POST.get('title')
            activity.activity_type = request.POST.get('activity_type', 'field')
            activity.custom_activity_type = request.POST.get('custom_activity_type')
            activity.date = request.POST.get('date')
            activity.time = request.POST.get('time')
            activity.location = request.POST.get('location')
            activity.moderator = request.POST.get('moderator')
            activity.save()
            messages.success(request, f'Activité {activity.title} modifiée avec succès!')
            return redirect('evangelisation-list')
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    activity_choices = EvangelismActivity.ACTIVITY_TYPE_CHOICES
    return render(request, 'dashboard/evangelisation.html', {
        'activity': activity,
        'action': 'Modifier',
        'activity_choices': activity_choices
    })


@login_required
def evangelisation_delete(request, pk):
    """Supprimer une activité d'évangélisation"""
    activity = get_object_or_404(EvangelismActivity, pk=pk)
    
    if request.method == 'POST':
        activity.delete()
        messages.success(request, 'Activité supprimée avec succès!')
        return redirect('evangelisation-list')
    
    return render(request, 'dashboard/evangelisation.html', {'activity': activity, 'delete_confirm': True})


# ============================================================
# Formations - CRUD
# ============================================================

@login_required
def training_list(request):
    """Liste des formations"""
    trainings = TrainingEvent.objects.all().order_by('-date', '-time')
    return render(request, 'dashboard/trainings.html', {'trainings': trainings, 'view': 'trainings'})


@login_required
def training_create(request):
    """Créer une formation"""
    if request.method == 'POST':
        form = TrainingEventForm(request.POST)
        if form.is_valid():
            training = form.save()
            messages.success(request, f'Formation {training.title} créée avec succès!')
            return redirect('training-list')
    else:
        form = TrainingEventForm()
    
    return render(request, 'dashboard/trainings.html', {'form': form, 'action': 'Créer', 'view': 'form'})


@login_required
def training_edit(request, pk):
    """Modifier une formation"""
    training = get_object_or_404(TrainingEvent, pk=pk)
    
    if request.method == 'POST':
        form = TrainingEventForm(request.POST, instance=training)
        if form.is_valid():
            form.save()
            messages.success(request, f'Formation {training.title} modifiée avec succès!')
            return redirect('training-list')
    else:
        form = TrainingEventForm(instance=training)
    
    return render(request, 'dashboard/trainings.html', {
        'form': form,
        'training': training,
        'action': 'Modifier',
        'view': 'form'
    })


@login_required
def training_delete(request, pk):
    """Supprimer une formation"""
    training = get_object_or_404(TrainingEvent, pk=pk)
    
    if request.method == 'POST':
        training.delete()
        messages.success(request, 'Formation supprimée avec succès!')
        return redirect('training-list')
    
    return render(request, 'dashboard/trainings.html', {'training': training, 'view': 'delete_confirm'})


# ============================================================
# Mariages - CRUD
# ============================================================

@login_required
def marriage_list(request):
    """Liste des registres de mariage"""
    marriages = MarriageRecord.objects.all().select_related('groom', 'bride').order_by('-planned_date')
    return render(request, 'dashboard/mariage.html', {'marriages': marriages})


@login_required
def marriage_detail(request, pk):
    """Détail d'un registre de mariage"""
    marriage = get_object_or_404(MarriageRecord, pk=pk)
    return render(request, 'dashboard/mariage.html', {'marriage': marriage, 'view': 'detail'})


@login_required
def marriage_create(request):
    """Créer un registre de mariage"""
    if request.method == 'POST':
        try:
            marriage = MarriageRecord.objects.create(
                groom_id=request.POST.get('groom') or None,
                bride_id=request.POST.get('bride') or None,
                groom_full_name=request.POST.get('groom_full_name'),
                bride_full_name=request.POST.get('bride_full_name'),
                groom_birth_date=request.POST.get('groom_birth_date') or None,
                groom_birth_place=request.POST.get('groom_birth_place'),
                groom_nationality=request.POST.get('groom_nationality'),
                bride_birth_date=request.POST.get('bride_birth_date') or None,
                bride_birth_place=request.POST.get('bride_birth_place'),
                bride_nationality=request.POST.get('bride_nationality'),
                godfather_full_name=request.POST.get('godfather_full_name'),
                godfather_nationality=request.POST.get('godfather_nationality'),
                godmother_full_name=request.POST.get('godmother_full_name'),
                godmother_nationality=request.POST.get('godmother_nationality'),
                planned_date=request.POST.get('planned_date'),
                planned_time=request.POST.get('planned_time'),
                location=request.POST.get('location'),
                dowry_paid=request.POST.get('dowry_paid') == 'on',
                civil_verified=request.POST.get('civil_verified') == 'on',
                prenuptial_tests=request.POST.get('prenuptial_tests') == 'on',
                approved=request.POST.get('approved') == 'on',
                created_by=request.user
            )
            messages.success(request, 'Registre de mariage créé avec succès!')
            return redirect('marriage-detail', pk=marriage.pk)
        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {str(e)}')
    
    members = Member.objects.filter(is_active=True)
    return render(request, 'dashboard/mariage.html', {'members': members, 'action': 'Créer', 'view': 'form'})


@login_required
def marriage_edit(request, pk):
    """Modifier un registre de mariage"""
    marriage = get_object_or_404(MarriageRecord, pk=pk)
    
    if request.method == 'POST':
        try:
            marriage.groom_id = request.POST.get('groom') or None
            marriage.bride_id = request.POST.get('bride') or None
            marriage.groom_full_name = request.POST.get('groom_full_name')
            marriage.bride_full_name = request.POST.get('bride_full_name')
            marriage.groom_birth_date = request.POST.get('groom_birth_date') or None
            marriage.groom_birth_place = request.POST.get('groom_birth_place')
            marriage.groom_nationality = request.POST.get('groom_nationality')
            marriage.bride_birth_date = request.POST.get('bride_birth_date') or None
            marriage.bride_birth_place = request.POST.get('bride_birth_place')
            marriage.bride_nationality = request.POST.get('bride_nationality')
            marriage.godfather_full_name = request.POST.get('godfather_full_name')
            marriage.godfather_nationality = request.POST.get('godfather_nationality')
            marriage.godmother_full_name = request.POST.get('godmother_full_name')
            marriage.godmother_nationality = request.POST.get('godmother_nationality')
            marriage.planned_date = request.POST.get('planned_date')
            marriage.planned_time = request.POST.get('planned_time')
            marriage.location = request.POST.get('location')
            marriage.dowry_paid = request.POST.get('dowry_paid') == 'on'
            marriage.civil_verified = request.POST.get('civil_verified') == 'on'
            marriage.prenuptial_tests = request.POST.get('prenuptial_tests') == 'on'
            marriage.approved = request.POST.get('approved') == 'on'
            marriage.save()
            messages.success(request, 'Registre de mariage modifié avec succès!')
            return redirect('marriage-detail', pk=marriage.pk)
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    members = Member.objects.filter(is_active=True)
    return render(request, 'dashboard/mariage.html', {
        'marriage': marriage,
        'members': members,
        'action': 'Modifier',
        'view': 'form'
    })


@login_required
def marriage_delete(request, pk):
    """Supprimer un registre de mariage"""
    marriage = get_object_or_404(MarriageRecord, pk=pk)
    
    if request.method == 'POST':
        marriage.delete()
        messages.success(request, 'Registre de mariage supprimé avec succès!')
        return redirect('marriage-list')
    
    return render(request, 'dashboard/mariage.html', {'marriage': marriage, 'delete_confirm': True})


# ============================================================
# Documents - CRUD
# ============================================================

def format_file_size(size):
    """Formate la taille d'un fichier"""
    if size < 1024:
        return f"{size} o"
    elif size < 1024 * 1024:
        return f"{size / 1024:.1f} Ko"
    elif size < 1024 * 1024 * 1024:
        return f"{size / (1024 * 1024):.1f} Mo"
    else:
        return f"{size / (1024 * 1024 * 1024):.1f} Go"


@login_required
def document_list(request):
    """Liste des documents avec recherche et filtres"""
    documents = Document.objects.all().order_by('-uploaded_at')

    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        documents = documents.filter(title__icontains=search_query)

    # Filtre par type de document
    doc_type_filter = request.GET.get('doc_type', '')
    if doc_type_filter:
        documents = documents.filter(document_type=doc_type_filter)

    # Filtre par format de fichier
    type_filter = request.GET.get('type', '')
    if type_filter:
        if type_filter == 'pdf':
            documents = documents.filter(file__iendswith='.pdf')
        elif type_filter == 'word':
            documents = documents.filter(file__iregex=r'\.(docx?)$')
        elif type_filter == 'excel':
            documents = documents.filter(file__iregex=r'\.(xlsx?)$')
        elif type_filter == 'image':
            documents = documents.filter(file__iregex=r'\.(jpe?g|png|gif|bmp)$')

    # Statistiques
    pdf_count = documents.filter(file__iendswith='.pdf').count()
    office_count = documents.filter(file__iregex=r'\.(docx?|xlsx?)$').count()
    image_count = documents.filter(file__iregex=r'\.(jpe?g|png|gif|bmp)$').count()

    # Ajouter taille formatée à chaque document
    for doc in documents:
        doc.file_size_formatted = format_file_size(doc.file.size) if doc.file else "0 o"

    return render(request, 'dashboard/documents.html', {
        'documents': documents,
        'view': 'list',
        'search_query': search_query,
        'doc_type_filter': doc_type_filter,
        'type_filter': type_filter,
        'pdf_count': pdf_count,
        'office_count': office_count,
        'image_count': image_count
    })


@login_required
def document_create(request):
    """Uploader un document"""
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.uploaded_by = request.user
            document.save()
            messages.success(request, f'Document {document.title} uploadé avec succès!')
            return redirect('document-list')
    else:
        form = DocumentForm()
    
    return render(request, 'dashboard/documents.html', {'form': form, 'action': 'Uploader', 'view': 'form'})


@login_required
def document_detail(request, pk):
    """Détail d'un document"""
    document = get_object_or_404(Document, pk=pk)
    document.file_size_formatted = format_file_size(document.file.size) if document.file else "0 o"
    return render(request, 'dashboard/documents.html', {'document': document, 'view': 'detail'})


@login_required
def document_edit(request, pk):
    """Modifier un document (métadonnées uniquement, pas le fichier)"""
    document = get_object_or_404(Document, pk=pk)

    if request.method == 'POST':
        form = DocumentEditForm(request.POST, instance=document)
        if form.is_valid():
            form.save()
            messages.success(request, f'Document {document.title} modifié avec succès!')
            return redirect('document-list')
    else:
        form = DocumentEditForm(instance=document)

    document.file_size_formatted = format_file_size(document.file.size) if document.file else "0 o"
    return render(request, 'dashboard/documents.html', {
        'form': form,
        'document': document,
        'action': 'Modifier',
        'view': 'form'
    })


@login_required
def document_delete(request, pk):
    """Supprimer un document"""
    document = get_object_or_404(Document, pk=pk)
    
    if request.method == 'POST':
        document.delete()
        messages.success(request, 'Document supprimé avec succès!')
        return redirect('document-list')
    
    return render(request, 'dashboard/documents.html', {'document': document, 'view': 'delete_confirm'})


# ============================================================
# Baptêmes - CRUD
# ============================================================

@login_required
def baptism_list(request):
    """Liste des événements de baptême"""
    baptisms = BaptismEvent.objects.all().select_related('event').prefetch_related('candidates').order_by('-event__date')
    return render(request, 'dashboard/baptism.html', {'baptisms': baptisms, 'view': 'list'})


@login_required
def baptism_detail(request, pk):
    """Détail d'un événement de baptême"""
    baptism = get_object_or_404(BaptismEvent.objects.select_related('event').prefetch_related('candidates'), pk=pk)
    return render(request, 'dashboard/baptism.html', {'baptism': baptism, 'view': 'detail'})


@login_required
def baptism_create(request):
    """Créer un événement de baptême"""
    if request.method == 'POST':
        # Créer l'événement associé
        event_form = EventForm(request.POST, prefix='event')
        if event_form.is_valid():
            event = event_form.save()
            baptism = BaptismEvent.objects.create(event=event, created_by=request.user)
            messages.success(request, 'Événement de baptême créé avec succès!')
            return redirect('baptism-list')
    else:
        event_form = EventForm(prefix='event')

    return render(request, 'dashboard/baptism.html', {
        'event_form': event_form,
        'view': 'form',
        'action': 'Créer'
    })


@login_required
def baptism_edit(request, pk):
    """Modifier un événement de baptême"""
    baptism = get_object_or_404(BaptismEvent.objects.select_related('event'), pk=pk)

    if request.method == 'POST':
        event_form = EventForm(request.POST, instance=baptism.event, prefix='event')
        if event_form.is_valid():
            event_form.save()
            messages.success(request, 'Événement de baptême modifié avec succès!')
            return redirect('baptism-list')
    else:
        event_form = EventForm(instance=baptism.event, prefix='event')

    return render(request, 'dashboard/baptism.html', {
        'event_form': event_form,
        'baptism': baptism,
        'view': 'form',
        'action': 'Modifier'
    })


@login_required
def baptism_delete(request, pk):
    """Supprimer un événement de baptême"""
    baptism = get_object_or_404(BaptismEvent, pk=pk)

    if request.method == 'POST':
        baptism.event.delete()  # Supprime aussi l'événement associé
        messages.success(request, 'Événement de baptême supprimé avec succès!')
        return redirect('baptism-list')

    return render(request, 'dashboard/baptism.html', {
        'baptism': baptism,
        'view': 'delete_confirm'
    })


@login_required
def baptism_candidate_add(request, pk):
    """Ajouter un candidat au baptême"""
    baptism = get_object_or_404(BaptismEvent, pk=pk)

    if request.method == 'POST':
        form = BaptismCandidateForm(request.POST, request.FILES)
        if form.is_valid():
            candidate = form.save(commit=False)
            candidate.baptism_event = baptism
            candidate.save()
            messages.success(request, 'Candidat ajouté avec succès!')
            return redirect('baptism-detail', pk=pk)
    else:
        form = BaptismCandidateForm()

    return render(request, 'dashboard/baptism.html', {
        'form': form,
        'baptism': baptism,
        'view': 'candidate_form'
    })


# ============================================================
# Contact Admin (Gestion des messages reçus)
# ============================================================

@login_required
def contact_admin_list(request):
    """Liste des messages de contact reçus"""
    contacts = Contact.objects.all().order_by('-created_at')
    status_filter = request.GET.get('status', '')

    if status_filter:
        contacts = contacts.filter(status=status_filter)

    return render(request, 'dashboard/contact_admin.html', {
        'contacts': contacts,
        'status_filter': status_filter,
        'view': 'list'
    })


@login_required
def contact_admin_detail(request, pk):
    """Détail d'un message de contact"""
    contact = get_object_or_404(Contact, pk=pk)

    # Marquer comme lu automatiquement
    if contact.status == 'new':
        contact.status = 'read'
        contact.save(update_fields=['status'])

    return render(request, 'dashboard/contact_admin.html', {
        'contact': contact,
        'view': 'detail'
    })


@login_required
def contact_mark_read(request, pk):
    """Marquer un message comme lu"""
    contact = get_object_or_404(Contact, pk=pk)
    contact.status = 'read'
    contact.save(update_fields=['status'])
    messages.success(request, 'Message marqué comme lu.')
    return redirect('contact-admin-list')


@login_required
def contact_archive(request, pk):
    """Archiver un message"""
    contact = get_object_or_404(Contact, pk=pk)
    contact.status = 'archived'
    contact.save(update_fields=['status'])
    messages.success(request, 'Message archivé.')
    return redirect('contact-admin-list')


# ============================================================
# Audit Logs (Logs système)
# ============================================================

@login_required
@admin_required
def audit_log_list(request):
    """Liste des logs système avec filtres avancés"""
    logs = AuditLogEntry.objects.all().select_related('actor').order_by('-created_at')

    # Filtres
    action_filter = request.GET.get('action', '')
    model_filter = request.GET.get('model', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if action_filter:
        logs = logs.filter(action=action_filter)
    if model_filter:
        logs = logs.filter(model__icontains=model_filter)
    if user_filter:
        logs = logs.filter(
            actor__username__icontains=user_filter
        ) | logs.filter(
            actor__first_name__icontains=user_filter
        ) | logs.filter(
            actor__last_name__icontains=user_filter
        )
    if date_from:
        logs = logs.filter(created_at__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__lte=date_to + ' 23:59:59')

    # Pagination manuelle - limiter à 500
    logs = logs[:500]

    # Statistiques
    total_logs = AuditLogEntry.objects.count()
    create_count = AuditLogEntry.objects.filter(action='create').count()
    update_count = AuditLogEntry.objects.filter(action='update').count()
    delete_count = AuditLogEntry.objects.filter(action='delete').count()

    # Liste des modèles uniques pour le filtre
    unique_models = AuditLogEntry.objects.values_list('model', flat=True).distinct().order_by('model')

    return render(request, 'dashboard/audit_logs.html', {
        'logs': logs,
        'action_filter': action_filter,
        'model_filter': model_filter,
        'user_filter': user_filter,
        'date_from': date_from,
        'date_to': date_to,
        'action_choices': AuditLogEntry.ACTION_CHOICES,
        'unique_models': unique_models,
        'total_logs': total_logs,
        'create_count': create_count,
        'update_count': update_count,
        'delete_count': delete_count
    })


# ============================================================
# Demandes d'approbation (ApprovalRequest)
# ============================================================

@login_required
def approval_request_list(request):
    """Liste des demandes d'approbation"""
    # Les demandes créées par l'utilisateur
    my_requests = ApprovalRequest.objects.filter(requested_by=request.user).order_by('-created_at')

    # Les demandes en attente de décision (pour les admins)
    if request.user.is_superuser or request.user.role in ['admin', 'pastor', 'super_admin']:
        pending_requests = ApprovalRequest.objects.filter(status='pending').order_by('-created_at')
    else:
        pending_requests = ApprovalRequest.objects.none()

    # Filtres
    status_filter = request.GET.get('status', '')
    if status_filter:
        my_requests = my_requests.filter(status=status_filter)

    return render(request, 'dashboard/approval_requests.html', {
        'my_requests': my_requests,
        'pending_requests': pending_requests,
        'status_filter': status_filter,
        'view': 'list'
    })


@login_required
def approval_request_detail(request, pk):
    """Détail d'une demande d'approbation"""
    approval_request = get_object_or_404(ApprovalRequest, pk=pk)

    # Vérifier les permissions
    is_owner = approval_request.requested_by == request.user
    can_decide = request.user.is_superuser or request.user.role in ['admin', 'pastor', 'super_admin']

    return render(request, 'dashboard/approval_requests.html', {
        'approval_request': approval_request,
        'is_owner': is_owner,
        'can_decide': can_decide,
        'view': 'detail'
    })


@login_required
@admin_required
def approval_request_approve(request, pk):
    """Approuver une demande"""
    approval_request = get_object_or_404(ApprovalRequest, pk=pk)

    if request.method == 'POST':
        approval_request.status = 'approved'
        approval_request.decided_by = request.user
        approval_request.decided_at = timezone.now()
        approval_request.save()

        # Créer une notification pour le demandeur
        Notification.objects.create(
            title='Demande approuvée',
            message=f'Votre demande de {approval_request.get_action_display()} pour {approval_request.object_repr} a été approuvée.',
            recipient=approval_request.requested_by
        )

        messages.success(request, 'Demande approuvée avec succès!')
        return redirect('approval-request-list')

    return redirect('approval-request-detail', pk=pk)


@login_required
@admin_required
def approval_request_reject(request, pk):
    """Rejeter une demande"""
    approval_request = get_object_or_404(ApprovalRequest, pk=pk)

    if request.method == 'POST':
        reason = request.POST.get('rejection_reason', '')
        approval_request.status = 'rejected'
        approval_request.decided_by = request.user
        approval_request.decided_at = timezone.now()
        approval_request.rejection_reason = reason
        approval_request.save()

        # Créer une notification pour le demandeur
        Notification.objects.create(
            title='Demande rejetée',
            message=f'Votre demande de {approval_request.get_action_display()} pour {approval_request.object_repr} a été rejetée. Raison: {reason}',
            recipient=approval_request.requested_by
        )

        messages.success(request, 'Demande rejetée.')
        return redirect('approval-request-list')

    return redirect('approval-request-detail', pk=pk)


# ============================================================
# Notifications
# ============================================================

@login_required
def notification_list(request):
    """Liste des notifications de l'utilisateur"""
    notifications = Notification.objects.filter(recipient=request.user).order_by('-created_at')

    # Statistiques
    unread_count = notifications.filter(is_read=False).count()
    total_count = notifications.count()

    # Filtre lu/non lu
    filter_type = request.GET.get('filter', '')
    if filter_type == 'unread':
        notifications = notifications.filter(is_read=False)
    elif filter_type == 'read':
        notifications = notifications.filter(is_read=True)

    return render(request, 'dashboard/notifications.html', {
        'notifications': notifications,
        'unread_count': unread_count,
        'total_count': total_count,
        'filter_type': filter_type,
        'view': 'list'
    })


@login_required
def notification_mark_read(request, pk):
    """Marquer une notification comme lue"""
    notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
    notification.is_read = True
    notification.save(update_fields=['is_read'])
    messages.success(request, 'Notification marquée comme lue.')
    return redirect('notification-list')


@login_required
def notification_mark_all_read(request):
    """Marquer toutes les notifications comme lues"""
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    messages.success(request, 'Toutes les notifications ont été marquées comme lues.')
    return redirect('notification-list')


@login_required
def notification_delete(request, pk):
    """Supprimer une notification"""
    notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
    notification.delete()
    messages.success(request, 'Notification supprimée.')
    return redirect('notification-list')


# ============================================================
# Compte utilisateur
# ============================================================

@login_required
def account(request):
    """Profil utilisateur complet"""
    user = request.user
    
    # Récupérer les informations du membre associé (si existe)
    try:
        member = Member.objects.select_related('user').get(user=user)
    except Member.DoesNotExist:
        member = None
    
    # Préparer les infos du membre pour l'affichage
    member_info = {}
    if member:
        member_info = {
            'Numéro de membre': member.member_number,
            'Date de naissance': member.birth_date.strftime('%d/%m/%Y') if member.birth_date else None,
            'Lieu de naissance': member.place_of_birth,
            'Genre': member.get_gender_display(),
            'Nationalité': member.nationality,
            'État civil': member.get_marital_status_display(),
            'Profession': member.occupation,
            'Fonction à l\'église': member.church_position,
            'Département': member.department.name if member.department else None,
            'Ministère': member.ministry.name if member.ministry else None,
            'Groupe de maison': member.home_group.name if member.home_group else None,
            'Date de baptême': member.baptism_date.strftime('%d/%m/%Y') if member.baptism_date else None,
            'Statut': 'Actif' if member.is_active else 'Inactif',
        }
    
    # Formulaires - Utiliser ProfileUpdateForm (sans role) pour son propre profil
    profile_form = ProfileUpdateForm(instance=user)
    password_form = PasswordChangeCustomForm(user=user)
    
    context = {
        'user': user,
        'member': member,
        'member_info': member_info,
        'profile_form': profile_form,
        'password_form': password_form,
        'active_page': 'account',
    }
    return render(request, 'dashboard/account.html', context)


@login_required
def account_edit(request):
    """Modifier le profil utilisateur"""
    user = request.user
    
    if request.method == 'POST':
        # Déterminer quel formulaire a été soumis
        form_type = request.POST.get('form_type', 'profile')
        
        if form_type == 'profile':
            form = ProfileUpdateForm(request.POST, request.FILES, instance=user)
            if form.is_valid():
                form.save()
                messages.success(request, 'Profil modifié avec succès!')
                return redirect('account')
            else:
                # Recharger la page avec les erreurs
                try:
                    member = Member.objects.select_related('user').get(user=user)
                except Member.DoesNotExist:
                    member = None
                
                # Préparer member_info aussi pour le cas d'erreur
                member_info = {}
                if member:
                    member_info = {
                        'Numéro de membre': member.member_number,
                        'Date de naissance': member.birth_date.strftime('%d/%m/%Y') if member.birth_date else None,
                        'Lieu de naissance': member.place_of_birth,
                        'Genre': member.get_gender_display(),
                        'Nationalité': member.nationality,
                        'État civil': member.get_marital_status_display(),
                        'Profession': member.occupation,
                        'Fonction à l\'église': member.church_position,
                        'Département': member.department.name if member.department else None,
                        'Ministère': member.ministry.name if member.ministry else None,
                        'Groupe de maison': member.home_group.name if member.home_group else None,
                        'Date de baptême': member.baptism_date.strftime('%d/%m/%Y') if member.baptism_date else None,
                        'Statut': 'Actif' if member.is_active else 'Inactif',
                    }
                
                context = {
                    'user': user,
                    'member': member,
                    'member_info': member_info,
                    'profile_form': form,
                    'password_form': PasswordChangeCustomForm(user=user),
                    'active_page': 'account',
                }
                return render(request, 'dashboard/account.html', context)

        elif form_type == 'password':
            password_form = PasswordChangeCustomForm(user=user, data=request.POST)
            if password_form.is_valid():
                password_form.save()
                messages.success(request, 'Mot de passe changé avec succès! Veuillez vous reconnecter.')
                return redirect('logout')
            else:
                # Recharger la page avec les erreurs
                try:
                    member = Member.objects.select_related('user').get(user=user)
                except Member.DoesNotExist:
                    member = None
                
                member_info = {}
                if member:
                    member_info = {
                        'Numéro de membre': member.member_number,
                        'Date de naissance': member.birth_date.strftime('%d/%m/%Y') if member.birth_date else None,
                        'Lieu de naissance': member.place_of_birth,
                        'Genre': member.get_gender_display(),
                        'Nationalité': member.nationality,
                        'État civil': member.get_marital_status_display(),
                        'Profession': member.occupation,
                        'Fonction à l\'église': member.church_position,
                        'Département': member.department.name if member.department else None,
                        'Ministère': member.ministry.name if member.ministry else None,
                        'Groupe de maison': member.home_group.name if member.home_group else None,
                        'Date de baptême': member.baptism_date.strftime('%d/%m/%Y') if member.baptism_date else None,
                        'Statut': 'Actif' if member.is_active else 'Inactif',
                    }
                
                context = {
                    'user': user,
                    'member': member,
                    'member_info': member_info,
                    'profile_form': ProfileUpdateForm(instance=user),
                    'password_form': password_form,
                    'active_page': 'account',
                }
                return render(request, 'dashboard/account.html', context)
    
    return redirect('account')


# ============================================================
# Paramètres de l'église
# ============================================================

@login_required
def church_settings_view(request):
    """Vue pour gérer les paramètres de l'église"""
    settings = ChurchSettings.get_settings()

    if request.method == 'POST':
        form = ChurchSettingsForm(request.POST, request.FILES, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'Paramètres de l\'église mis à jour avec succès!')
            return redirect('church-settings')
        else:
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        form = ChurchSettingsForm(instance=settings)

    # Récupérer ou créer la biographie
    church_biography = ChurchBiography.objects.filter(is_active=True).first()

    # Récupérer les cultes
    church_services = ChurchService.objects.all().order_by('order', 'day', 'time')

    context = {
        'form': form,
        'settings': settings,
        'active_page': 'settings',
        'activities_preview': ChurchActivity.objects.filter(is_active=True).order_by('order', 'title')[:6],
        'church_biography': church_biography,
        'church_services': church_services,
    }
    return render(request, 'dashboard/church_settings.html', context)


@login_required
def church_biography_view(request):
    """Vue pour gérer la biographie de l'église"""
    biography = ChurchBiography.objects.filter(is_active=True).first()
    
    if not biography:
        biography = ChurchBiography.objects.create(
            title="Biographie de l'église",
            content="Bienvenue sur la page de notre église.",
            is_active=True
        )
    
    if request.method == 'POST':
        form = ChurchBiographyForm(request.POST, request.FILES, instance=biography)
        if form.is_valid():
            form.save()
            messages.success(request, 'Biographie mise à jour avec succès!')
            return redirect('church-biography')
        else:
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        form = ChurchBiographyForm(instance=biography)
    
    return render(request, 'dashboard/church_biography.html', {
        'form': form,
        'biography': biography,
        'active_page': 'settings',
    })


@login_required
def church_activities_view(request):
    """Vue pour gérer les activités de l'église"""
    activities = ChurchActivity.objects.all().order_by('order', 'title')
    return render(request, 'dashboard/church_activities.html', {
        'activities': activities,
        'active_page': 'settings',
    })


@login_required
def activity_create_view(request):
    """Vue pour créer une nouvelle activité"""
    if request.method == 'POST':
        form = ChurchActivityForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Activité créée avec succès!')
            return redirect('church-activities')
        else:
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        form = ChurchActivityForm()
    
    return render(request, 'dashboard/activity_form.html', {
        'form': form,
        'action': 'create',
        'active_page': 'settings',
    })


@login_required
def activity_edit_view(request, pk):
    """Vue pour modifier une activité"""
    activity = get_object_or_404(ChurchActivity, pk=pk)
    
    if request.method == 'POST':
        form = ChurchActivityForm(request.POST, instance=activity)
        if form.is_valid():
            form.save()
            messages.success(request, 'Activité mise à jour avec succès!')
            return redirect('church-activities')
        else:
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        form = ChurchActivityForm(instance=activity)
    
    return render(request, 'dashboard/activity_form.html', {
        'form': form,
        'activity': activity,
        'action': 'edit',
        'active_page': 'settings',
    })


@login_required
def activity_delete_view(request, pk):
    """Vue pour supprimer une activité"""
    activity = get_object_or_404(ChurchActivity, pk=pk)
    
    if request.method == 'POST':
        activity.delete()
        messages.success(request, 'Activité supprimée avec succès!')
        return redirect('church-activities')
    
    return render(request, 'dashboard/activity_confirm_delete.html', {
        'activity': activity,
        'active_page': 'settings',
    })


@login_required
def church_services_view(request):
    """Vue pour gérer les horaires des cultes"""
    services = ChurchService.objects.all().order_by('order', 'day', 'time')
    return render(request, 'dashboard/church_services.html', {
        'services': services,
        'active_page': 'settings',
    })


@login_required
def service_create_view(request):
    """Vue pour créer un nouveau culte"""
    if request.method == 'POST':
        form = ChurchServiceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horaire de culte créé avec succès!')
            return redirect('church-services')
        else:
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        form = ChurchServiceForm()
    
    return render(request, 'dashboard/service_form.html', {
        'form': form,
        'action': 'create',
        'active_page': 'settings',
    })


@login_required
def service_edit_view(request, pk):
    """Vue pour modifier un horaire de culte"""
    service = get_object_or_404(ChurchService, pk=pk)
    
    if request.method == 'POST':
        form = ChurchServiceForm(request.POST, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horaire de culte mis à jour avec succès!')
            return redirect('church-services')
        else:
            messages.error(request, 'Veuillez corriger les erreurs ci-dessous.')
    else:
        form = ChurchServiceForm(instance=service)
    
    return render(request, 'dashboard/service_form.html', {
        'form': form,
        'service': service,
        'action': 'edit',
        'active_page': 'settings',
    })


@login_required
def service_delete_view(request, pk):
    """Vue pour supprimer un horaire de culte"""
    service = get_object_or_404(ChurchService, pk=pk)
    
    if request.method == 'POST':
        service.delete()
        messages.success(request, 'Horaire de culte supprimé avec succès!')
        return redirect('church-services')
    
    return render(request, 'dashboard/service_confirm_delete.html', {
        'service': service,
        'active_page': 'settings',
    })
